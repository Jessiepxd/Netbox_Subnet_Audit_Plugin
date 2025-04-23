# ==================================================inspection====================
# DeviceLab - Commissioning Tool
# wxPython 4.1.0+ is required (www.wxpython.org)
# ======================================================================
import os                   # Cross-platform operating system interface.
import sys                  # Python system module
import time                 # Python timing
# import json                 # JSON formatting
import struct               # Binary data packing/unpacking.
import locale               # Set locale for wxPython.
import socket               # IP Address handling.
import typing               # Type checking.
# import textwrap             # Word-wrapping long strings.
# import datetime             # Date/Time handling.
import threading            # Multiple threads
import traceback            # Generate error messages.
import subprocess           # Used to run commands.
import collections          # Used for namedtuple and deque.
try:    # wxPython GUI Framework
    import wx
    # if wx.VERSION < (4, 1):
    #     raise ImportError("wxPython %s too old. Require 4.1.0+" % wx.version())
    import wx.adv           # wxPython - Advanced widgets (i.e. About Dialog).
    import wx.lib.masked    # wxPython - Masked controls.
    import wx.lib.mixins.inspection as WIT
    import wx.lib.agw.hypertreelist as HTL
    import images           # Application Images wxPython EmbeddedImages.
except ImportError as err:
    print(str(err))
    if 'libgtk-3' in str(err):
        # libgtk-3.so.0: cannot open shared object file: No such file or directory
        print("GTK3 libraries required.")
        print("\tsudo apt install libgtk-3-0")
    elif 'libnotify' in str(err):
        # libnotify.so.4: cannot open shared object file
        print("Desktop notifications library required.")
        print("\tsudo apt install libnotify4")
    elif 'libsdl' in str(err).lower():
        print("Simple DirectMedia Layer library required.")
        library = 'libsdl1.2-dev' if '1.2' in str(err) else 'libsdl2-2.0-0'
        print("\tsudo apt install %s" % library)
    else:
        # Assume wxPython itself is not installed.
        print("wxPython required: http://www.wxpython.org")
        print("python -m pip install wxpython")
        print("Linux wheels: https://extras.wxpython.org/wxPython4/extras/linux/gtk3/")
    sys.exit(1)
try:    # OpenPyXL for reading/writing .XLSX files.
    import openpyxl
except ImportError as err:
    raise ImportError('\n'.join((
        str(err), "Requires OpenPyXL: https://openpyxl.readthedocs.io/en/stable/",
        "\tpip install openpyxl")))
if not openpyxl.DEFUSEDXML:
    # OpenPyXL uses defusedxml to protect against XML attacks.
    raise ImportError('\n'.join((
        "Requires defusedxml: https://pypi.org/project/defusedxml/",
        "\tpip install defusedxml")))
if not openpyxl.LXML:
    # LXML uses efficient C libraries libxml2 and libxslt for speed.
    raise ImportError('\n'.join(("Requires lxml: https://lxml.de/",
                                 "\tpip install lxml")))
# Import models class __init__
import models.__info__ as models

# Application defines --------------------------------------------------
DEBUG = False               # Enable/disable the debug stderr/stdout window
APPNAME = "DeviceLab"
VENDORNAME = "Evertz"
COMPANYNAME = "Evertz Microsystems Ltd."
PRODUCTNAME = "Device Commissioning Tool"
COPYRIGHT = "2023 Evertz Microsystems Ltd."
VERSION = "0.1"             # Initial version/template.
VERSION = "0.2"             # Scanning subnets working.
VERSION = "1.0"             # Tested Scorpion Modules
VERSION = "1.1"             # Started implementing IPGs
VERSION = "1.2"             # Added Dialog after upgrade
VERSION = "1.3"             # Added Dialog colors and fixes
# Version information for Windows "VersionInfo" resource.
VERSIONPARTS = [int(part) for part in VERSION.split('.')]
VERSIONPARTS += [0] * max(0, 4 - len(VERSIONPARTS))
VERSIONINFO = {
    'fileMajor': VERSIONPARTS[0], 'fileMinor': VERSIONPARTS[1],
    'fileMicro': VERSIONPARTS[2], 'fileBuild': VERSIONPARTS[3],
    'prodMajor': VERSIONPARTS[0], 'prodMinor': VERSIONPARTS[1],
    'prodMicro': VERSIONPARTS[2], 'prodBuild': VERSIONPARTS[3],
    'companyName': COMPANYNAME, 'fileDescription': PRODUCTNAME,
    'fileVersion': VERSION, 'internalName': APPNAME,
    'legalCopyright': COPYRIGHT, 'originalFilename': APPNAME + '.exe',
    'productName': PRODUCTNAME, 'productVersion': VERSION,
    'comments': '', 'legalTrademarks': '', }

SUPPORTED_DEVICE_LIST = [
    "scorpion2", "scorpion4", "scorpion6", "scorpionx18", "scorpions18"
]

# Initialize wxPython Application -------------------------------------
DEBUG = False
args = [arg.lower() for arg in sys.argv]
if '/debug' in args or '-v' in args or '-d' in args:
    DEBUG = True
app = None
# If running, initialize the wxPython app immediately.
if __name__ == "__main__":
    # Creating the App may initialize the locale.
    # app = wx.App(DEBUG)
    app = WIT.InspectableApp(DEBUG)
    # Force the locale to the "compatible" 'C' locale
    locale.setlocale(locale.LC_ALL, 'C')
    # wxlocale = wx.Locale(wx.LANGUAGE_DEFAULT)
    # Used to place config in Windows registry
    app.SetAppName(APPNAME)
    app.SetVendorName(VENDORNAME)
    if tuple(int(v) for v in HTL.__version__.split('.')) < (1, 6):
        wx.MessageBox("Warning old hypertreelist.py v%s" % HTL.__version__,
                      caption="WX Library Warning", style=wx.ICON_EXCLAMATION)

# ---------------------------------------------------------------------
# Import third-party modules. These may have hidden dependencies (DLLs).
try:
    import ahttp                        # Evertz asynchronous HTTP operations.
    import asnmp                        # Evertz asynchronous SNMP operations.
    if not asnmp.AES or not asnmp.DES:
        raise ImportError("Requires pycryptodomex for SNMP v3 support")
    import anbt                         # Evertz asynchronous NetBIOS library.
    from snmpdata import ENTERPRISES    # { enterprise_num: (name, weblink) }
except ImportError as err:
    # Fail gracefully if there are missing OS dependencies.
    if app:
        wx.MessageBox("Error Loading\n%s" % str(err), caption="Startup Failed",
                      style=wx.OK | wx.CENTER | wx.ICON_ERROR)
    raise

# Helper functions ----------------------------------------------------
def copy_to_clipboard(text: str):
    """Copy the given text to the clipboard as a wx.TextDataObject."""
    if not wx.TheClipboard.IsOpened():
        success = wx.TheClipboard.Open()
    if success is True:
        data = wx.TextDataObject(text)
        success = wx.TheClipboard.SetData(data)
        wx.TheClipboard.Close()
    if success is False:
        wx.MessageBox("Failed to copy to the clipboard")


# Classes -------------------------------------------------------------
class InterfaceInfo(object):
    """Class to hold information about an interface on a device.

    self.if_num = The interface number assigned by its OperatingSystem.
    self.MAC = MAC address in colon-separated lower-case hex format.
    self.name = The interface name (possibly blank)
    self.IP = The assigned IP (possibly blank)
    """
    def __init__(self, if_num: int, mac_octets: str, name: str = ""):
        """Immediately parse the mac_str into colon-separated format."""
        self.if_num = if_num
        if not mac_octets or len(mac_octets) != 6:
            raise ValueError("Invalid MAC address string")
        if not isinstance(mac_octets, (bytes, bytearray)):
            mac_octets = mac_octets.encode('l1', errors="ignore")
        mac_bytes = struct.unpack('6B', mac_octets)
        self.MAC = ":".join("%02x" % octet for octet in mac_bytes)
        self.name = name
        self.IP = ""

    def __repr__(self) -> str:
        """InterfaceInfo(if_num=2, mac="00:02:c5:12:33:f3", name="eth0")"""
        return "%s(if_num=%s, mac=%s, name=%s)" % (
            self.__class__.__name__, self.if_num, self.MAC, self.name)
#end class InterfaceInfo(object)


class DeviceInfo(object):
    """Container class to hold info on discovered devices.

    self.IP = String IP address of device.
    self.name = Discovered name of the device.
    self.icon = Name of icon to use in Tree View.
    self.ok = True if device was contacted, False otherwise.
    self.model = Model object containing model found.
    self.current_version = Software 'major'.'minor'.'build'
    self.selected = If the device is selected
    self.GROUP = The group the device belongs to
    self.current_firmware = The current firmware for the device

    We probe a device for SNMP/LLDP, HTTP/WebEASY, and NetBIOS.
    The SNMP results go into self.snmp:
      self.snmp = {OID: value} and will be empty if the device did not
                  responds to SNMP packets.
        { sysDescr, sysName, sysObjectID, sysLocation }
    All HTTP results go into self.http:
      self.http = {
          'content': bytes,     # Raw bytes of the '/' path, if not WebEASY.
          'headers': dict       # Response headers as {name: value}.
          'status_code': int    # The HTTP response code for the '/' path.
          'version': str,       # WebEASY version, if found, as "v.#.#".
          'https': bool,        # If the HTTP server was HTTP or HTTPS.
          'cfg_json': dict,     # Reply from cfgJSON request (if any).
          'cfg_web': dict,      # Reply from cfgWeb request (if any).
          }

    From the above, we populate self.results with all processed
    information from all replies. It can contain any of the following:
      self.results = {
        'enterprise': int,      # From sysObjectID or SNMPv3 engineID.
        'vendor': str,          # From 'enterprise' number, if found.
        'netbios': str,         # Possibly from NetBIOS query.
        'card_name': str,       # From cardName OID or WebEASY 1@s.
        'serial': str,          # From boardSerialNumber OID or WebEasy 8@s.
        'mac': str,             # The best-guess MAC address for this device.
        'alias': str,           # Card alias name given by user.
        'major': int,           # FW version softwareRevisionMajor OID or WebEasy 3@i.
        'minor': int,           # FW version softwareRevisionMinor OID or WebEASY 4@i.
        'build': int,           # FW version softwarePointReleaseNumber or WebEASY 6@s.
        'board_name': str,      # boardName OID or WebEASY 9@s.
        'board_rev': str,       # boardRevision OID or Webeasy 10@s
        'board_build': int,     # boardBuild OID or Webeasy 11@i
        'http_name': str,       # Best-guess name of HTTP server or `None`.
        }

    Additionally we try to get interface information. This is obtained
    via SNMP IF-MIB (if supported) or SNMPv3 EngineID (if reported).
      self.interfaces = { MAC: (interface_name, IP_address) }
    Note it could be { MAC: (None, None) } if no information was
    extracted for a specific MAC address.
    """
    # Reserved VarIDs.
    VARIDS = ('1@s',    # Product Name (e.g. "570IPG-X19")
              '2@s',    # Creation Date (e.g. "20180207")
              '3@i',    # Revision Major (e.g. "1")
              '4@i',    # Revision Minor (e.g. "0")
              '6@s',    # Build Number (e.g. "0452-App F")
              '8@s',    # Serial Number (e.g. "7678130045")
              '9@s',    # Name PRODUCT (e.g. "570IPG")
              '10@s',   # Revision BRDREV (e.g. "A")
              '11@i',)  # Build Number BRDBLD (e.g. "2")
    # SNMP v2 management OIDs to fetch.
    OIDS = (asnmp.sysDescr, asnmp.sysName, asnmp.sysObjectID, asnmp.sysLocation)
    # SNMP v2 Evertz common MIB OIDs to fetch.
    MINICARD_VALUES = ('cardName', 'boardSerialNumber', 'softwareRevisionMajor',
                       'softwareRevisionMinor', 'softwarePointReleaseNumber',
                       'boardName', 'boardRevision', 'hardwareBuildNumber')
    MINICARD = {name: asnmp.MINICARD[name] for name in MINICARD_VALUES}
    # POLL_TIMEOUT = 20   # Timeout for a device to poll
    POLL_TIMEOUT = 15   # Timeout for a device to poll
    SNMP_RETRIES = 1    # Default number of retries for SNMP operations.
    SNMP_TIMEOUT = 2.0  # Default timeout per retry of SNMP operations.
    MAX_SNMP = 1        # Maximum number of simultaneous SNMP operations.

    def __init__(self, IP, snmp_thread, http_thread, nbt_thread):
        # Convert IP address to 32-bit representation.
        try:
            self.IP32 = struct.unpack("!L", socket.inet_aton(IP))[0]
        except (socket.error, struct.error):
            raise ValueError("Invalid address %s" % repr(IP))
        # Convert IP back to string representation.
        self.IP = '.'.join(str((self.IP32 >> i) & 0xFF) for i in (24, 16, 8, 0))
        self.GROUP = ''               # The device group it belongs to
        self.icon = "unknown"           # Case-sensitive
        self.ok = False                 # True if refresh succeeded.
        self.http_thread: ahttp.HttpThread = http_thread
        self.snmp_thread: asnmp.SnmpThread = snmp_thread
        self.nbt_thread: anbt.NbtThread = nbt_thread
        self.model = None
        self.https = False          # Start with regular HTTP.
        self.started = False        # False, or timestamp of when probe started.
        self.selected = False       # If device is selected for upgrade/license
        self.elapsed = 0            # Time in seconds probe has taken.
        self.percent = 0            # Percentage probe is done.
        self.http = {}              # HTTP results from probe.
        self.snmp = {}              # SNMP results from probe.
        self.cfgweb = False
        self.cfgjson = False
        self.results = {}           # Specific information from probe.
        self.interfaces = {}        # { MAC: (interface_name, ip_address) }
        self.requests = {}          # Ongoing operations.
        self.profile = {}
        self.nbt = []               # NetBIOS results from probe
        self.TIMEOUT = ((self.SNMP_RETRIES + 1) * self.SNMP_TIMEOUT) + 0.5
        self.current_firmware = ""  # Software Version 'major.minor.build'

    def http_get(self, path, params=None, headers={}, safe=',',
                 allow_redirects=False, max_redirects=2, queue=None,
                 callback=None, timeout=4.0, https=None):
        """Accessor method for non-blocking HTTP or HTTPS gets.

        Equivalent to calling:
        self.http_thread.get('http://%s%s' % (self.IP,path), block=False)

        Will use 'http' if https is False, or 'https' if https is True.
        If https is not specified, uses self.https to determine schema.
        By default allow_redirects is False so won't follow redirects.
        Returns the non-blocking operation object.
        """
        if https is None:
            https = self.https
        schema = 'http' if https is False else 'https'
        return self.http_thread.get("%s://%s%s" % (schema, self.IP, path),
                                    params=params, headers=headers, block=False,
                                    safe=safe, allow_redirects=allow_redirects,
                                    max_redirects=max_redirects, queue=queue,
                                    callback=callback, timeout=timeout)

    def snmp_get(self, requests, community='public', op_retries=None,
                 op_timeout=None, timeout=4.0, queue=None, callback=None):
        """Delegator method for non-blocking SNMP multiple gets.

        Equivalent to calling:
        self.snmp_thread.batchGet(self.IP, requests, block=False)

        'requests' should be a list of operations. Each operation can
        contain one or more OIDs (up to 1024). The maximum number of
        simultaneous pending SNMP operations is set by self.MAX_SNMP.
        The default number of retries and timeout per attempt are set by
        self.SNMP_RETRIES and self.SNMP_TIMEOUT, if not specified.
        """
        if op_retries is None:
            op_retries = self.SNMP_RETRIES
        if op_timeout is None:
            op_timeout = self.SNMP_TIMEOUT
        return self.snmp_thread.batchGet(self.IP, requests, community=community,
                                         op_retries=op_retries, timeout=timeout,
                                         op_timeout=op_timeout, queue=queue,
                                         callback=callback, block=False,
                                         max_pending=self.MAX_SNMP)

    def snmp_walk(self, base_oid, queue=None, max_results=None,
                  community='public', version=0, bulk=False, max_repetitions=1,
                  timeout=10.0, op_retries=1, op_timeout=2.0, callback=None):
        '''Delegator method for non-blocking SNMP walks.

        Equivalent to calling:
        self.snmp_thread.walk(self.IP, base_oid, block=False)

        Will perform repeated GetNext or GetBulk (bulk=True) operations
        on the agent until either:
            1. the returned OIDs no longer branch from the 'base_oid'.
            2. 'max_results' (if specified) is returned.
            3. The timeout elapses.
        By default 'max_results' is None to disable that check.

        The results of the walk are added to the queue as 2-tuples of
        (branch_oid, value) where 'branch_oid' is a tuple containing the
        last OID value numers, not including the 'base_oid' portion.
        If a queue is not specified, one is created.

        Returns a WalkOperation() object.
        '''
        return self.snmp_thread.walk(
            self.IP, base_oid, queue=queue, max_results=max_results,
            community=community, version=version, bulk=bulk,
            max_repetitions=max_repetitions, timeout=timeout, block=False,
            op_retries=op_retries, op_timeout=op_timeout, callback=callback)

    def get_cfgweb(self, parameters, path='/cgi-bin/cfgweb', timeout=4.0,
                   https=None, queue=None, callback=None):
        '''Delegator for non-blocking http_thread.get_cfgweb(self.IP).

        If https is not specified, will use self.https.
        Always non-blocking (block=False)
        '''
        if https is None:
            https = self.https
        return self.http_thread.get_cfgweb(
            self.IP, parameters, path=path, timeout=timeout, https=https,
            queue=queue, callback=callback, block=False)

    def get_cfgjson(self, parameters, path='/cgi-bin/cfgjsonrpc',
                    version='2.0', jsonId=None, https=None, queue=None,
                    callback=None, timeout=4.0):
        '''Delegator for self.http_thread.get_cfgjson(self.IP)

        If https is not specified, will use self.https.
        Always non-blocking (block=False)
        The operation.result will contain { varId: value }.
        '''
        if https is None:
            https = self.https
        return self.http_thread.get_cfgjson(
            self.IP, parameters, path=path, version=version, jsonId=jsonId,
            https=https, queue=queue, callback=callback, timeout=timeout,
            block=False)

    def probe(self):
        """Initiate a probe of this device for SNMP/HTTP/NBT probes."""
        # Don't do anything if already refreshing (start timestamp set).
        if self.started:
            return
        self.started = time.time()
        self.percent = 0
        self.ok = False
        self.http.clear()
        self.snmp.clear()
        self.interfaces.clear()
        self.nbt = []

        # Start the http operation to get WebEASY version.
        self.requests['httpOp'] = self.http_get(
            '/version.txt', allow_redirects=True, timeout=4.5)
        # Start the first SNMP operation.
        self.requests['snmpOp'] = self.snmp_thread.get(
            self.IP, asnmp.sysDescr, version=0, timeout=self.SNMP_TIMEOUT,
            retries=self.SNMP_RETRIES, block=False)
        # TODO Start first SNMPv3 operation
        self.requests['snmpv3Op'] = self.snmp_thread.get(
            self.IP, asnmp.sysObjectID, version=3, userName='evertz', block=False)
        # Start NetBIOS query to device.
        if self.nbt_thread and self.nbt_thread.is_alive() is True:
            self.requests['nbtOp'] = self.nbt_thread.query(self.IP, block=False)

    def poll(self):
        """Check ongoing SNMP/HTTP/NBT operations for results.

        Returns percent complete."""
        # Return 100% if no active poll ongoing (maybe we finished already).
        if not self.started:
            return 100
        # Process ongoing operations and grab results.
        snmpOp = self.process_snmp(self.requests.pop('snmpOp', None))
        snmpv3Op = self.process_snmpv3(self.requests.pop('snmpv3Op', None))
        nbtOp = self.process_nbt(self.requests.pop('nbtOp', None))
        httpOp = self.process_http(self.requests.pop('httpOp', None))
        # Push back into self.requests dictionary (if new operations created).
        self.requests['snmpOp'] = snmpOp
        self.requests['snmpv3Op'] = snmpv3Op
        self.requests['nbtOp'] = nbtOp
        self.requests['httpOp'] = httpOp
        # Search for a model if SNMP and HTTP operations are complete.
        model = None
        # if httpOp is None and snmpOp is None:
        #     self.model = models.find(self)
        #     if model:
        #         self.poll_finished(model=model)
        #         return 100
        # Check for poll timeout.
        self.elapsed = abs(time.time() - self.started)
        if self.elapsed > self.POLL_TIMEOUT:
            # We took too long. Force everything to stop.
            print("Device %s timed out %ssec." % (self.IP, self.POLL_TIMEOUT))
            self.poll_finished(model=None)
            self.percent = 100
        # Check if all operations are finished.
        if not any((snmpOp, snmpv3Op, nbtOp, httpOp)):
            self.poll_finished(model)
            self.percent = 100
        return self.percent

    def process_http(self, httpOp: ahttp.HttpOperation):
        """Process current http operation (might be None).

        Returns the operation if not complete yet or None if finished.
        Will return a new HTTP operation on successful response.
        """
        # Is the http request complete?
        if httpOp and httpOp.finished is True:
            self.profile.setdefault(httpOp.url, []).append(abs(time.time() - httpOp.started))
            URL = httpOp.req_url.lower()
            if URL.endswith('version.txt'):     # WebEASY version -----
                webVer = None
                if httpOp.ok is True:
                    # Device is running WebEASY. Extract version.
                    content = httpOp.content.decode('utf8', errors='replace')
                    for line in content.splitlines():
                        if line.startswith('V') and line[1:4].isdigit() is True:
                            webVer = "v.%c.%c" % (line[1], line[2])
                    # Set self.https to True if we got redirected to https.
                    if httpOp.url.lower().startswith('https'):
                        self.https = True
                if webVer:
                    # Put into self.http results.
                    self.http['version'] = webVer
                    # Start cfgWeb/cfgJSON requests
                    version = float(webVer[2:])
##                    print("%s %.1fs - Got webEasy version %r, https=%s" %
##                          (self.IP, elapsed, version, self.https))
                    if version < 1.5:
                        # WebEASY v1.4 or older mostly uses cfgWeb API.
                        httpOp = self.get_cfgweb(self.VARIDS)
                        self.cfgweb = True
                    else:
                        # WebEASY v1.5 supports cfgJSON API and maybe alias.
                        httpOp = self.get_cfgjson(self.VARIDS + ('9150@s',))
                        self.cfgjson = True
                    if httpOp.ok is True:
                        params = {varid: value
                                  for varid, value in httpOp.result.items()
                                  if varid and value}
                        # Populate self.results
                        self.extract_results(params)
                else:
                    # No webeasy. Try to fetch default URL.
                    httpOp = self.http_get(
                        '/', allow_redirects=True, timeout=self.TIMEOUT)
            elif URL.endswith('/'):     # Generic HTTP -----------------
                # Record results of HTTP operation (if any) and end probe.
                #  print("%s %.1fs - Generic HTTP: ok=%s, error=%s, status=%s" %
                #        (self.IP, elapsed, httpOp.ok, httpOp.error, httpOp.status))
                if httpOp.content:
                    self.http['status'] = httpOp.status_code    # e.g. 200
                    self.http['headers'] = httpOp.headers       # {name: value}
                    self.http['content'] = httpOp.content       # Raw bytes
                self.percent += 25
                httpOp = None
            elif URL.endswith('cfgjsonrpc'):    # cfgJSON --------------
                # Extract cfgJSON parameters, if any.
                params = {}
                # print("%s %.1fs - cfgJSON complete, ok=%s, error=%s, %d params" %
                #       (self.IP, elapsed, httpOp.ok, httpOp.error, len(httpOp.result)))
                if httpOp.ok is True:
                    params = {varid: value
                              for varid, value in httpOp.result.items()
                              if varid and value is not None or value != ""}
                self.http.setdefault('cfg_json', {}).update(params)
                # self.results=['']
                # Populate self.results
                self.extract_results(params)
                if not params and 'cfg_web' not in self.http:
                    # cfgJSON failed. Try cfgWeb probe.
                    # print("%s cfgJSON failed, attempting cfgWeb" % self.IP)
                    httpOp = self.get_cfgweb(self.VARIDS)
                else:
                    # We got our cfgJSON so we're done!
                    self.percent += 25
                    httpOp = None
            elif URL.endswith('cfgweb'):    # cfgWeb -------------------
                # Extract cfgWeb parameters, if any.
                params = {}
                if httpOp.ok is True:
                    params = {varid: value
                              for varid, value in httpOp.result.items()
                              if varid and value}
##                    print("%s %.1fs - cfgWeb complete, %d params" %
##                          (self.IP, elapsed, 0 if not params else len(params)))
                self.http.setdefault('cfg_web', {}).update(params)
                # Populate self.results
                self.extract_results(params)
                if not params and 'cfg_json' not in self.http:
                    # cfgWeb failed. Try cfgJSON probe.
                    #print("%s CfgWeb failed. Starting cfgJSON" % self.IP)
                    httpOp = self.get_cfgjson(self.VARIDS)
                elif params and '9150@s' not in params:
                    # cfgWeb alias hack - Request separately in case it fails.
                    #print("%s requesting cfgWeb alias" % self.IP)
                    httpOp = self.get_cfgweb(('9150@s',))
                else:
                    # We got our cfgWeb parameters so we're done!
                    self.percent += 25
                    httpOp = None
            else:
                print("*** Unknown request URL: %r" % URL)
                self.percent += 25
                httpOp = None
        return httpOp

    def process_snmp(self, snmpOp):
        """Process current SNMP operation (might be None).

        Returns either the same operation (if not finished) or None if
        SNMP probe is completed, or a new operation if continuing.
        """
        IF_PHYS_MIB = "1.3.6.1.2.1.2.2.1.6"     # ifPhysAddress table
        IF_DESC_MIB = "1.3.6.1.2.1.2.2.1.2"     # ifDescr table
        LLDP_MIB = "1.0.8802.1.1.2.1.3.7.1"     # lldpLocPortEntry table.
        if snmpOp and snmpOp.finished is True:
            self.profile.setdefault(snmpOp.__class__.__name__, []).append(abs(time.time() - snmpOp.started))
            if isinstance(snmpOp, asnmp.SNMPOperation):
                # debug
                if self.IP == '172.16.199.10':
                    print(f"{self.IP}: snmp: {snmpOp.error}, oid: {snmpOp.oids}")
                if snmpOp.oids:
                    # Got a response for sysDescr, device running an SNMP agent.
                    self.snmp.update(snmpOp.oids)
                    # Schedule GET for next critial OID.
                    missing = [oid for oid in self.OIDS if oid not in self.snmp]
                    if missing:
                        # Get the next OID in self.OIDS.
                        snmpOp = self.snmp_thread.get(
                            self.IP, missing[0], version=0,
                            timeout=self.SNMP_TIMEOUT,
                            retries=self.SNMP_RETRIES, block=False)
                    else:
                        # All critial OIDs probed. Process sysObjectID.
                        self.extract_results(self.snmp)
                        # Try to get Evetz MINICARD.
                        snmpOp = self.snmp_thread.batchGet(
                            self.IP, [[OID] for OID in self.MINICARD.values()],
                            op_retries=self.SNMP_RETRIES, timeout=self.TIMEOUT,
                            op_timeout=self.SNMP_TIMEOUT, max_pending=1,
                            block=False)
                else:
                    # No response for a critical OID. Stop all SNMP operations.
                    # print("DEBUG: No OIDS found. IP:%s ok: %s, error: %s" % (self.IP, snmpOp.ok, snmpOp.error))
                    self.percent += 25
                    snmpOp = None
            if isinstance(snmpOp, asnmp.GetBatchOperation):
                # Batch get result for Evertz MINICARD OIDs.
                if snmpOp.oids:
                    # Process the SNMP MINICARD OIDs into self.results.
                    self.extract_results(self.snmp)
                # All SNMP OIDs have been probed. Try walking IF-MIB.
                snmpOp = self.snmp_walk(IF_PHYS_MIB, bulk=True)
            elif isinstance(snmpOp, asnmp.WalkOperation):
                # Reply to a walk operation. Was it successful?
                if snmpOp.ok is True and snmpOp.queue.qsize():
                    # We got results! Process them into self.interfaces.
                    results = [snmpOp.queue.get() for _ in range(snmpOp.queue.qsize())]
                    # results = [(tuple(oids), value), ...]
                    if snmpOp.base_oid == IF_PHYS_MIB:
                        # Got a walk reply to get Physical Addresses
                        results = {suffix[0]: value for suffix, value in results
                                   if len(suffix) == 1 and value}
                        for if_num, octets in results.items():
                            try:
                                interface_info = InterfaceInfo(if_num, octets)
                            except ValueError:
                                pass
                            else:
                                self.interfaces[if_num] = interface_info
                        # Got phyical, send walk for IF_DESC_MIB
                        snmpOp = self.snmp_walk(IF_DESC_MIB, bulk=True)
                    elif snmpOp.base_oid == IF_DESC_MIB:
                        # Got a walk reply to interface names.
                        results = {suffix[0]: value for suffix, value in results
                                   if len(suffix) == 1 and value}
                        for if_num, if_name in results.items():
                            if if_num in self.interfaces:
                                self.interfaces[if_num].name = if_name
                                # Add eth 0 mac address to the results
                                if if_name == 'eth0':
                                    self.results['mac'] = self.interfaces.get(if_num).MAC
                        # Done! The IF-MIB gave us everything, no need for LLDP.
                        self.percent += 25
                        snmpOp = None
                    elif snmpOp.base_oid == LLDP_MIB:
                        # Got a walk reply for LLDP-MIB. Parse the results.
                        lldp = {}   # { port_num: {table_oid: value} }
                        for suffix, value in results:
                            if len(suffix) == 2:
                                # suffix = (table_oid, port_num)
                                table_oid, port_num = suffix
                                lldp_data = lldp.setdefault(port_num, {})
                                lldp_data[table_oid] = value
                        # Scan through lldp dictionary and record MAC addresses.
                        for port_num, lldp_data in lldp.items():
                            subtype = lldp_data.get(2)      # .2 = PortIdSubtype
                            if subtype == 3:    # macAddress (3)
                                octets = lldp_data.get(3)   # .3 = PortId
                                name = lldp_data.get(4)     # .4 = Description
                                try:
                                    self.interfaces[port_num] = InterfaceInfo(
                                        port_num, octets, name=name)
                                    self.results['mac'] = self.interfaces[port_num].MAC
                                except ValueError:
                                    pass
                        # All done, no more requests to make.
                        self.percent += 25
                        snmpOp = None
                else:
                    # No response to SNMP walk.
                    if snmpOp.base_oid == IF_PHYS_MIB:
                        # IF-MIB not try LLDP-MIB.
                        snmpOp = self.snmp_walk(LLDP_MIB, bulk=True)
                    else:
                        # Walks failed. Stop SNMP probe.
                        self.percent += 25
                        snmpOp = None
        return snmpOp

    def process_snmpv3(self, snmpv3Op):
        """ Process current SNMPv3 operation (might be None).

        Returns either the same operation (if not finished) or None if
        SNMP probe is completed, or a new operation if continuing.

        """
        # If device supports snmpv3 look at the engine ID and flag that snmpv3 is supported.
        if snmpv3Op and snmpv3Op.finished is True:
            self.profile.setdefault('SNMPv3', []).append(abs(time.time() - snmpv3Op.started))
            # print(f"DEBUG: EngineID:{snmpv3Op.engineId}")
            if snmpv3Op.engineId and len(snmpv3Op.engineId) >= 5:
                # Extract enterprise ID from engineID:
                enterprise_num = struct.unpack('!L', snmpv3Op.engineId[:4])[0] & 0x7fffffff
                self.results['enterprise'] = enterprise_num
                vendor, website = ENTERPRISES.get(enterprise_num, (None, None))
                self.results['vendor'] = vendor
                if snmpv3Op.engineId[4] == b'\x03':
                    # Score! engineId contains a MAC address!
                    mac_bytes = snmpv3Op.engineId[5:11]
                    if len(mac_bytes) == 6:
                        self.interfaces[0] = InterfaceInfo(0, mac_bytes)
                        print("Got %s SNMPv3 enterprise %s (%s), MAC: %s" % (self.IP, enterprise_num, vendor, self.interfaces[0].MAC))
                        # set the mac address if its not set already
                        if not self.results.get('mac'):
                            self.results['mac'] = self.interfaces[0].MAC
            # Stop SNMPv3 probe.
            self.percent += 25
            snmpv3Op = None
        return snmpv3Op

    def process_nbt(self, nbtOp: anbt.NbtOperation):
        """Process current NetBIOS operation (might be None).

        Returns either the same operation (if not finished) or None if
        NetBIOS query is finished.
        """
        # Check up on NetBIOS query.
        if nbtOp and nbtOp.finished is True:
            self.profile.setdefault('NBT', []).append(abs(time.time() - nbtOp.started))
            # nbtOp.reply.names is a list of 2-tuples of (name, flags).
            if nbtOp.ok is True and nbtOp.reply:
                # Sort names so lower value of flags is at start of list.
                names = sorted(nbtOp.reply.names, key=lambda n: n[1])
                # Filter names to only workstation IDs (ends with \x00).
                names = [name[:15].strip() for name, flags in names
                         if name.endswith('\x00')]
                if names:
                    self.results['netbios'] = names[0]
            # Stop NBT probe.
            self.percent += 25
            nbtOp = None
        return nbtOp

    def poll_finished(self, model=None):
        """Stop the poll and set the model (if found)."""
        self.started = False
        self.requests.clear()
        #
        # print("%s finished in %.1f sec." % (self.IP, self.elapsed))
        # print("DEBUG:poll_finished, IP:%s, results:%s, snmp: %s, http %s" % (self.IP, self.results, self.snmp, self.http))
        # print("%s finished results %.1s sec." % (self.IP, self.results))
        # if self.elapsed > 10:
        #     for name, times in self.profile.items():
        #         print("\t%s times: %s" % (name, ', '.join("%.1f" % t for t in times)))

        # Set the http_name in self.results if we got an HTTP response.
        if self.http.get('version'):
            # It's running WebEASY
            self.results['http_name'] = str(self.http['version'])
        elif self.http:
            # Try to get the HTML <title> element, otherwise Server header.
            http_name = self.http.get('headers', {}).get('Server', '')
            body = self.http.get('content', b'')
            title = body[body.find(b'<title>'):body.find(b'</title>')][7:]
            http_name = title if len(title) > len(http_name) else http_name
            if hasattr(http_name, 'decode'):
                http_name = http_name.decode('utf8', 'ignore')
            self.results['http_name'] = http_name
        else:
            # The HTTP title is not found
            http_name = None

        # Pick the best interface from self.interfaces, if we have any.
        sorted_interfaces = sorted(self.interfaces.items())
        if sorted_interfaces:
            self.results['interface'] = sorted_interfaces[0][1]
        # Get the model
        self.model = models.find(self)
        # Assign our icon.
        if self.model:
            self.icon = 'evertz'    # Known and modelled device
        elif self.results.get('enterprise', None) == 6827:  # Evertz
            self.icon = 'evertz'   # Evertz Microsystems Ltd.
        elif self.results.get('enterprise', None) == 8072:     # Net-SNMP
            self.icon = 'linux'     # Net-SNMP (server-based product).
        elif self.results.get('vendor', None):
            self.icon = 'gear'      # Other registered SNMP vendor.
        elif self.results.get('netbios'):
            self.icon = 'pc'        # Likely a PC, server or workstation.
        elif self.http:
            self.icon = 'web'       # Web-reply only.
        else:
            self.icon = 'unknown'
        if self.model or self.results:
            self.ok = True
        #     print("DEBUG:Poll_finished I AM OKAY! %s, ip:%s, resp:%s " % (self.ok, self.IP, self.results))
        # else:
        #     print("DEBUG:Poll_finished DeviceInfo object is not set to ok... %s, ip:%s, resp:%s " % (self.ok, self.IP, self.results))

    def extract_results(self, params: dict):
        """Translate parameters from VarIDs to self.results dictionary.

            params: { VarID: value } or { OID: value }

            'card_name': str,       # From cardName OID or WebEASY 1@s.
            'serial': str,          # From boardSerialNumber OID or WebEasy 8@s.
            'major': int,           # FW version softwareRevisionMajor OID or WebEasy 3@i.
            'minor': int,           # FW version softwareRevisionMinor OID or WebEASY 4@i.
            'build': int,           # FW version softwarePointReleaseNumber or WebEASY 6@s.
            'board_name': str,      # boardName OID or WebEASY 9@s.
            'board_rev': str,       # boardRevision OID or Webeasy 10@s
            'board_build': int,     # boardbuild OID or Webeasy 11@i
            'alias': str,           # Card alias from WebEASY 9150@s.
        """
        # print("DEBUG:extract_results, IP:%s, params: %s, results:%s, snmp: %s, http %s" % (self.IP, params, self.results, self.snmp, self.http))
        # Process sysObjectID to results['enterprise'] and ['vendor'].
        sysObjectID = params.get(asnmp.sysObjectID)
        if sysObjectID and isinstance(sysObjectID, str):
            # Split ObjectID into enterprises component.
            if sysObjectID.startswith('1.3.6.1.4.1.'):
                oid, dot, rest = sysObjectID[12:].partition('.')
                if oid.isdigit() is True:
                    vendor, weblink = ENTERPRISES.get(int(oid), (None, None))
                    self.results['enterprise'] = int(oid)
                    if vendor:
                        self.results['vendor'] = vendor
        # 'card_name' is either SNMP OID or WebEASY VarID '1@s'.
        card_name = params.get('1@s', params.get(self.MINICARD['cardName']))
        if card_name:
            self.results['card_name'] = card_name
        serial = params.get('8@s', params.get(self.MINICARD['boardSerialNumber']))
        if serial:
            self.results['serial'] = str(serial)
        major = params.get('3@i', params.get(self.MINICARD['softwareRevisionMajor']))
        if major is not None and major != "":
            try:
                self.results['major'] = int(major)
            except (TypeError, ValueError):
                pass
        minor = params.get('4@i', params.get(self.MINICARD['softwareRevisionMinor']))
        if minor is not None and minor != "":
            try:
                self.results['minor'] = int(minor)
            except (TypeError, ValueError):
                pass
        build = params.get('6@s', params.get(self.MINICARD['softwarePointReleaseNumber']))
        if build:
            self.results['build'] = str(build)
        board_name = params.get('9@s', params.get(self.MINICARD['boardName']))
        if board_name:
            self.results['board_name'] = str(board_name)
        board_rev = params.get('10@s', params.get(self.MINICARD['boardRevision']))
        if board_rev:
            self.results['board_rev'] = str(board_rev)
        board_build = params.get('11@i', params.get(self.MINICARD['hardwareBuildNumber']))
        if board_build is not None and board_build != "":
            try:
                self.results['board_build'] = int(board_build)
            except (TypeError, ValueError):
                pass
        alias = params.get('9150@s')
        if alias:
            self.results['alias'] = str(alias)
        # Setup the firmware version
        self.current_firmware = '%s.%s.%s' % (self.results.get('major', ''),
                                              self.results.get('minor', ''),
                                              self.results.get('build', ''))
#end class DeviceInfo(object)

class ScanThread(threading.Thread):
    """A Thread to perform device discovery on a list of IPs"""
    def __init__(self, baseIP: int, count: int, name="Scan Thread", **kwargs):
        # Initialize Thread.
        super(ScanThread, self).__init__(name=name, **kwargs)
        self.startIP = baseIP & 0xffffffff
        self.endIP = (self.startIP + count) & 0xffffffff
        # Create new HTTP, SNMP, and NBT threads for this scan.
        self.http_thread = ahttp.start()
        self.snmp_thread = asnmp.start()
        self.nbt_thread = anbt.start()
        self.polling = []
        self.discovered = collections.deque()
        self.end_event = threading.Event()
        self.percent = 0
        self.status = "Initializing"

    def run(self):
        """This is the method that is run in a separate thread."""
        # 1) Create DeviceInfo object for each IP address to poll.
        # for IP32 in range(16, 18 + 1):
        for IP32 in range(self.startIP, self.endIP + 1):
            IP = socket.inet_ntoa(struct.pack("!L", IP32))
            deviceinfo = DeviceInfo(IP, snmp_thread=self.snmp_thread,
                                    http_thread=self.http_thread,
                                    nbt_thread=self.nbt_thread)
            deviceinfo.probe()
            self.polling.append(deviceinfo)
        # 2) Poll each device and add to self.discovered when complete.
        try:
            self.scan()
        except Exception as err:
            traceback.print_exc()
            self.status = "Exception %s" % str(err)
        else:
            self.status = "Aborted" if self.polling else "Finished"
        # 3) Stop all threads.
        self.http_thread.stop(block=False)
        self.snmp_thread.stop(block=False)
        self.nbt_thread.stop(block=False)

    def scan(self):
        """Scan all DeviceInfo objects in self.polling."""
        print("Scanning %s to %s (%s total)" % (self.polling[0].IP, self.polling[-1].IP, len(self.polling)))
        total_poll = len(self.polling)
        while self.polling and not self.end_event.is_set():
            self.status = "Scanning %s IPs" % len(self.polling)
            lowest_percent = 100
            for deviceinfo in tuple(self.polling):
                # Poll the device and update lowest_percent.
                percent = deviceinfo.poll()
                lowest_percent = min(percent, lowest_percent)
                if percent >= 100:
                    # Remove from self.device and add to self.discovered.
                    self.polling.remove(deviceinfo)
                    if deviceinfo.ok is True:
                        self.discovered.append(deviceinfo)
                    # else:
                    #     print("DEBUG:SCAN DeviceInfo object is not set to ok... %s, ip:%s, resp:%s " % (deviceinfo.ok, deviceinfo.IP, deviceinfo.results))
            # self.percent = lowest_percent
            self.percent = round((total_poll - len(self.polling)) * 100 / total_poll)
            time.sleep(0.010)

    def stop(self, block=True):
        """Stop the thread, optionally blocking until fully stopped."""
        self.end_event.set()
        if block is True:
            self.join()
#end class ScanThread(threading.Thread)


class ScanList(wx.ListCtrl):
    """List control to show discovered device information.

    self.devices = List of DeviceInfo class to show.
    """
    ICONS = ['ok_green', 'ok_yellow', 'warning', 'error', 'spin0', 'spin1',
             'spin2', 'spin3', 'spin4', 'spin5', 'spin6', 'spin7', 'progress0',
             'progress1', 'progress2', 'progress3', 'progress4', 'progress5',
             'progress6', 'progress7', 'progress8', 'wait0', 'wait1', 'wait2',
             'wait3', 'wait4', 'wait5', 'wait6', 'wait7', 'evertz', 'unknown',
             'linux', 'gear', 'pc', 'web']
    COLUMNS = (("IP", wx.LIST_FORMAT_LEFT, 3.6),
               ("Card Name", wx.LIST_FORMAT_CENTRE, 3.6),
               ("Serial", wx.LIST_FORMAT_CENTRE, 2.0),
               ("NetBIOS", wx.LIST_FORMAT_CENTRE, 2.5),
               ("HTTP", wx.LIST_FORMAT_CENTRE, 4.0),
               ("SNMP", wx.LIST_FORMAT_CENTRE, 5.5),
               ("sysName", wx.LIST_FORMAT_CENTRE, 4.5),
               ("sysDescr", wx.LIST_FORMAT_CENTRE, 4.5),
               ("Interfaces", wx.LIST_FORMAT_CENTRE, 5.0), )
    CHECKMARK = '\u2714'
    CROSSMARK = '\u274c'

    def __init__(self, parent, devices: dict, menu=None, style=0):
        # Call the original constructor to do its job. Force style
        style |= wx.LC_REPORT | wx.LC_VIRTUAL
        wx.ListCtrl.__init__(self, parent, style=style)
        # Assign our parent book and figure out our page number.
        self.parent: wx.Window = parent             # Parent window
        self.menu: wx.Menu = menu                   # Popup context menu.
        # This is all devices in our list control.
        self.devices: typing.Dict[int, DeviceInfo] = devices
        # Add standard columns
        UNIT = max(int(wx.SystemSettings.GetMetric(wx.SYS_ICON_X) * 1.125), 36)
        for title, alignment, width in self.COLUMNS:
            self.AppendColumn(title, format=alignment, width=int(width * UNIT))
        # Create imagelist for icons.
        self.icons = {}
        imagelist = wx.ImageList(16, 16)
        for name in self.ICONS:
            self.icons[name] = imagelist.Add(images.catalog[name].GetIcon())
        self.AssignImageList(imagelist, wx.IMAGE_LIST_SMALL)
        self.Bind(wx.EVT_RIGHT_DOWN, self.OnRightDown)

    def OnGetItemText(self, item_num, column):
        """Return text for the given item number (row) and column number."""
        if item_num >= 0 and item_num < len(self.devices):
            model_obj: DeviceInfo = self.devices[item_num]
            if column == 0:         # Device IP Address
                return model_obj.IP
            if column == 1:         # Card Name
                return model_obj.results.get("card_name", 'Not Found')
            if column == 2:         # Serial Number
                return model_obj.results.get("serial", 'Not Found')
            if column == 3:         # NetBIOS query result
                return model_obj.results.get("netbios", '-')
            if column == 4:         # Summary of HTTP
                http_name = model_obj.results.get("http_name", '-')
                return http_name if http_name else '-'
            if column == 5:         # Summary of SNMP
                return "%s (%s)" % (model_obj.results.get("vendor"),
                                    model_obj.results.get("enterprise"))
            if column == 6:         # sysName
                return model_obj.snmp.get(asnmp.sysName, '-')
            if column == 7:         # sysDescr
                return model_obj.snmp.get(asnmp.sysDescr, '-')
            if column == 8:         # Interfaces
                # Show the lowest interface MAC/name and total number.
                interface_info = model_obj.results.get('interface')
                if interface_info:
                    return "Total %s, %s=%s" % (len(model_obj.interfaces),
                                                interface_info.name,
                                                interface_info.MAC)
                else:
                    # No interface information returned.
                    return '-'
        # Unhandled row/column. Return blank string.
        return ""

    def OnGetItemImage(self, item_num):
        """Return an icon for the given source on the given index."""
        if item_num >= 0 and item_num < len(self.devices):
            device_info: DeviceInfo = self.devices[item_num]
            return self.icons.get(device_info.icon, wx.NO_IMAGE)
        return wx.NO_IMAGE

    def OnGetItemAttr(self, item_num):
        """Return the wx.ItemAttr() object to set appearance of the item."""
        # Return None to use default visual attributes.
        return None

    def GetSelections(self):
        """Return a list of all selected items in the list control.

        Returns a list of item indexes that are selected. Can be empty.
        """
        items = []
        index = self.GetFirstSelected()
        while index >= 0 and index < len(self.devices):
            items.append(index)
            index = self.GetNextSelected(index)
        return items

    def OnRightDown(self, event: wx.MouseEvent):
        """On Right click display pop-up menu, if one was defined."""
        if not self.menu:
            return
        item = wx.NOT_FOUND
        focused_item = self.GetFocusedItem()
        # Was this event from a mouse click or menu key press?
        if event.x >= 0 and event.y >= 0:
            point = event.GetPosition()
            item, flags = self.HitTest(point)   # Mouse click.
        elif focused_item >= 0 and focused_item < len(self.devices):
            item = focused_item
            point = self.GetItemRect(item).GetBottomLeft()
        # If right-click item not selected, clear selection then select it.
        selections = self.GetSelections()
        if item >= 0 and item < len(self.devices):
            if item not in selections:
                for selected_item in selections:
                    self.Select(selected_item, False)
            self.Select(item)
        self.SetFocus()
        self.PopupMenu(self.menu, point)
#end class ScanList(wx.ListCtrl)


class ScanDialog(wx.Dialog):
    """Dialog to allow scanning subnets for device discovery.

    self.scan_list is a wx.ListCtrl for each device discovered.
    The user data for each item contains the DeviceInfo object.
    """
    def __init__(self, parent: wx.Panel, wxconfig: wx.ConfigBase):
        wx.Dialog.__init__(self, parent, title="Scan Subnets for Devices",
                           style=wx.DEFAULT_DIALOG_STYLE | wx.RESIZE_BORDER, size=(1500, 2000))
        # self.SetSize((1000,1000))
        self.SetMinSize((-1, 700))
        # self.Refresh()
        self.parent: wx.Panel = parent
        self.wxconfig: wx.ConfigBase = wxconfig
        self.devices: typing.List[DeviceInfo] = []
        # Initialize UI -----------------------------------------------
        UNIT = max(int(wx.SystemSettings.GetMetric(wx.SYS_ICON_X) * 1.125), 36)
        ACVR = wx.ALIGN_CENTER_VERTICAL | wx.RIGHT
        main = wx.BoxSizer(orient=wx.VERTICAL)
        row = wx.BoxSizer(orient=wx.HORIZONTAL)
        label = wx.StaticText(self, label="Subnet IP")
        row.Add(label, flag=ACVR, border=UNIT // 6)
        self.subnet_combo = wx.ComboBox(self, style=wx.TE_PROCESS_ENTER)
        row.Add(self.subnet_combo, flag=ACVR, border=UNIT // 2)
        self.scan_button = wx.Button(self, label="&Scan")
        self.scan_button.SetToolTip("Scan this subnet")
        row.Add(self.scan_button, flag=ACVR)
        row.AddStretchSpacer()
        self.scan_text = wx.StaticText(self, label="", style=wx.ALIGN_CENTER)
        row.Add(self.scan_text)
        row.AddStretchSpacer()
        main.Add(row, flag=wx.EXPAND | wx.ALL, border=UNIT // 6)
        self.gauge = wx.Gauge(self, size=((-1, 3)))
        main.Add(self.gauge, flag=wx.EXPAND)
        self.scan_list = ScanList(self, self.devices)
        main.Add(self.scan_list, proportion=1, flag=wx.EXPAND | wx.ALL, border=UNIT // 6)
        row = wx.BoxSizer(orient=wx.HORIZONTAL)
        self.add_button = wx.Button(self, label="&Add Device")
        self.add_button.SetToolTip("Add the currently selected device")
        row.Add(self.add_button, flag=ACVR, border=UNIT // 6)
        self.add_all_button = wx.Button(self, label="Add Al&l")
        self.add_all_button.SetToolTip("Add all supported devices in the list")
        row.Add(self.add_all_button, flag=ACVR, border=UNIT // 6)
        row.AddStretchSpacer()
        self.ok_button = wx.Button(self, id=wx.ID_OK)
        row.Add(self.ok_button, flag=ACVR)
        main.Add(row, flag=wx.EXPAND | wx.ALL, border=UNIT // 6)
        self.SetSizer(main)
        self.Fit()
        # Retrieve past subnet values and populate combobox.
        self.history = list(filter(None, (
            self.wxconfig.Read('/subnets/%s' % num, '') for num in range(10))))
        if self.history:
            self.subnet_combo.SetItems(self.history[1:])
            self.subnet_combo.SetValue(self.history[0])
        # Bind events.
        self.subnet_combo.Bind(wx.EVT_TEXT_ENTER, self.OnScan)
        self.scan_button.Bind(wx.EVT_BUTTON, self.OnScan)
        self.add_button.Bind(wx.EVT_BUTTON, self.OnAdd)
        self.add_all_button.Bind(wx.EVT_BUTTON, self.OnAddAll)
        # self.ok_button.Bind(wx.EVT_BUTTON, self.OnOkay)
        #self.SetAffirmativeId(wx.ID_OK)

    def OnScan(self, event: wx.CommandEvent):
        """Start scanning the given subnet."""
        # Notify the user that we have started scanning
        self.scan_text.SetLabelText("Started Scan. Please wait around 20 seconds..")
        # Set the gauge to a default value
        self.gauge.SetValue(20)
        self.Refresh()
        self.Update()
        # Make sure a valid subnet has been entered into the subnet combobox.
        subnet: str = self.subnet_combo.GetValue()
        if subnet.count('.') < 3:
            subnet += '.0'
        if subnet and not subnet[-1].isdigit():
            subnet += '0'
        try:
            startIP32 = struct.unpack('!L', socket.inet_aton(subnet))[0]
            startIP32 = startIP32 & 0xffffff00
        except Exception as err:
            traceback.print_exc()
            wx.MessageBox("Invalid Subnet IP: '%s'\n%s" % (subnet, err),
                          caption="Bad Subnet", style=wx.ICON_ERROR | wx.OK)
            return
        # Push back into combobox and update history.
        subnetIP = socket.inet_ntoa(struct.pack('!L', startIP32))
        if self.subnet_combo.GetValue() != subnetIP:
            self.subnet_combo.SetValue(subnetIP)
        if subnetIP not in self.history:
            self.history.insert(0, subnetIP)
            if len(self.history) > 10:
                self.history.pop()
            for num, historyIP in enumerate(self.history):
                self.wxconfig.Write('/subnets/%s' % num, historyIP)
        # Clear self.devices, which is also referenced in scan_list.
        self.devices.clear()
        self.scan_list.DeleteAllItems()
        self.scan_list.Update()
        # Create the scan thread, and start it.
        thread = ScanThread(startIP32, count=255)
        # thread = ScanThread(startIP32, count=120)
        thread.start()
        # Monitor thread and populate self.devices.
        while thread.is_alive() is True:
            # proceed, skip = self.progress.Update(thread.percent)
            # if proceed is False:
            #     break
            if self.gauge.GetValue() != thread.percent:
                self.gauge.SetValue(thread.percent)
                self.gauge.Update()
            # Populate self.scan_list as devices are discovered.
            new_count = 0
            while thread.discovered:
                deviceinfo = thread.discovered.popleft()
                self.devices.append(deviceinfo)
                new_count += 1
            # Update the virtual list control with the number of devices.
            if new_count:
                self.scan_list.SetItemCount(len(self.devices))
                # Sort devices by numerical IP address
                self.devices.sort(key=lambda item: item.IP32)
                self.scan_list.Update()
                wx.SafeYield()
            # Go ahead and process events for screen updates.
            time.sleep(0.3)
        thread.stop()
        # Notify the user that we have finished scanning
        self.scan_text.SetLabelText("Finished Scan")
        self.gauge.SetValue(0)
        self.Refresh()

        # Send the device list to the appPanel
        self.parent.devices = self.devices
        # print("DEBUG: devices: %s" % thread.discovered)
        # print("DEBUG: list devices: %s" % self.devices )
        # for device_info in self.devices:
        #     print("%s - %s" % (device_info.IP, device_info.results))

    def OnAdd(self, event: wx.CommandEvent):
        """Add the currently selected device."""
        device_indexes = self.scan_list.GetSelections()

        print(f"device_indexes: {device_indexes}")
        for selected_index in device_indexes:
            if selected_index < 0 or selected_index >= len(self.devices):
                wx.MessageBox("No device selected", caption="Cannot Add")
                return
            else:
                device = self.devices[selected_index]
                # print(f"DEBUG: Index:{index}, device:{device}, results:{device.results}")
                self.parent.device_tree.AddDevice(device)
            # self.parent.added_devices.append(device)

    def OnAddAll(self, event: wx.CommandEvent):
        """Add all supported devices in self.devices."""
        for device in self.devices:
            self.parent.device_tree.AddDevice(device)
#end class SubnetDialog(wx.Dialog)


class DeviceTree(HTL.HyperTreeList):
    """Class to show all active devices and their status and data.

    """
    LABEL = "Unmatched"     # TopBook label
    DATETIME_FORMAT = "%a %b%d %I:%M:%S%p"
    COLUMNS = (('IP Address', 5.0, wx.ALIGN_LEFT),
               ('Supported', 2.0, wx.ALIGN_CENTER),
               ('Description', 4.0, wx.ALIGN_CENTER),
               ('Serial #', 2.0, wx.ALIGN_CENTER),
               ('MAC Address', 3.0, wx.ALIGN_CENTER),
               ('Alias', 3.0, wx.ALIGN_CENTER),
               ('Current Firmware', 4.0, wx.ALIGN_LEFT))
    # Icons to display in tree (lower-case) { name: (good_icon, bad_icon) }.
    ICONS = {
        "unknown": (images.unknown_g.GetBitmap, images.unknown_b.GetBitmap),
        "warning": (images.warning.GetBitmap, images.warning.GetBitmap),
        "error": (images.error.GetBitmap, images.error.GetBitmap),
        "info": (images.info.GetBitmap, images.info.GetBitmap),
        "subnet": (images.LAN_16.GetBitmap, images.LAN_16.GetBitmap),
        "evertz": (images.Evertz_g.GetBitmap, images.Evertz_b.GetBitmap),
        "gear": (images.gear_g.GetBitmap, images.gear_b.GetBitmap),
        "broom": (images.broom.GetBitmap, images.broom.GetBitmap),
        "linux": (images.Linux_g.GetBitmap, images.Linux_b.GetBitmap),
        "router": (images.router_g.GetBitmap, images.router_b.GetBitmap),
        "web": (images.web_g.GetBitmap, images.web_b.GetBitmap),
        "pc": (images.PC_g.GetBitmap, images.PC_b.GetBitmap),
        "spin0": (images.spin0.GetBitmap, images.spin0.GetBitmap),
        "spin1": (images.spin1.GetBitmap, images.spin1.GetBitmap),
        "spin2": (images.spin2.GetBitmap, images.spin2.GetBitmap),
        "spin3": (images.spin3.GetBitmap, images.spin3.GetBitmap),
        "spin4": (images.spin4.GetBitmap, images.spin4.GetBitmap),
        "spin5": (images.spin5.GetBitmap, images.spin5.GetBitmap),
        "spin6": (images.spin6.GetBitmap, images.spin6.GetBitmap),
        "spin7": (images.spin7.GetBitmap, images.spin7.GetBitmap),
        "wait0": (images.wait0.GetBitmap, images.wait0.GetBitmap),
        "wait1": (images.wait1.GetBitmap, images.wait1.GetBitmap),
        "wait2": (images.wait2.GetBitmap, images.wait2.GetBitmap),
        "wait3": (images.wait3.GetBitmap, images.wait3.GetBitmap),
        "wait4": (images.wait4.GetBitmap, images.wait4.GetBitmap),
        "wait5": (images.wait5.GetBitmap, images.wait5.GetBitmap),
        "wait6": (images.wait6.GetBitmap, images.wait6.GetBitmap),
        "wait7": (images.wait7.GetBitmap, images.wait7.GetBitmap), }

    def __init__(self, parent, menu=None, gauge=None, agwStyle=0):
        # BUG: Don't use HTL.TR_HIDE_ROOT as the paint is very slow (Dec 2022).
        HTL.HyperTreeList.__init__(self, parent, style=wx.TR_DEFAULT_STYLE,
                                   agwStyle=agwStyle | HTL.TR_DEFAULT_STYLE |
                                   HTL.TR_HAS_VARIABLE_ROW_HEIGHT |
                                   HTL.TR_ELLIPSIZE_LONG_ITEMS |
                                   HTL.TR_LINES_AT_ROOT | wx.TR_HAS_BUTTONS)
        self.parent = parent                # Parent usually a LabelBook.
        self.menu: wx.Menu = menu           # Optional context menu to show.
        self.gauge: wx.Gauge = gauge        # Optional wx.Gauge for progress.
        self.last_update = time.time()      # Last time we updated the tree.
        self.last_elapsed = 0               # Time taken to update tree last.
        self.parser = None                  # session.drain_unmatched.
        self.unmatched_count = 0            # Last update unmatched count.
        self.clusters = {}                  # {unmatched_id: TemplateCategory()}
        # Enable double-buffering as the draw is horrible without.
        self.SetBuffered(True)
        # Create imagelist
        self.iconIDs = {}
        imageList = wx.ImageList(16, 16)
        # Add self.ICONS to image list, keeping track of IDs.
        for name, (goodBitmap, badBitmap) in self.ICONS.items():
            goodID = imageList.Add(goodBitmap())
            badID = imageList.Add(badBitmap())
            self.iconIDs[name] = (goodID, badID)
        # Add provided rootBitmap to imagelist.
        report = wx.ArtProvider.GetBitmap(wx.ART_LIST_VIEW, size=(16, 16))
        self.iconIDs['root'] = imageList.Add(report)
        self.AssignImageList(imageList)

        # Setup columns.
        UNIT = max(int(wx.SystemSettings.GetMetric(wx.SYS_ICON_X) * 1.125), 36)
        for name, width, alignment in self.COLUMNS:
            self.AddColumn(name, width=int(width * UNIT), flag=alignment)
        self.SetMainColumn(0)
        # Add root node and start out expanded.
        root_id = self.AddRoot("Devices")
        self.root = root_id
        self.SetItemImage(root_id, self.iconIDs['root'])
        self.SetItemHasChildren(root_id, True)
        self.Expand(root_id)
        self.SelectItem(root_id)

        # Bind Events.
        self.Bind(wx.EVT_SIZE, self.OnTreeSize)
        self.Bind(wx.EVT_TREE_KEY_DOWN, self.OnTreeKeyDown)
        self.Bind(wx.EVT_TREE_ITEM_RIGHT_CLICK, self.OnTreeRightClick)

    def OnTreeSize(self, event):
        """Snoop size events to auto-size last column to max available width."""
        # Always allow event to propagate.
        event.Skip()
        # Get width and adjust for scrollbars.
        width = self.GetClientSize().width
        # Adust for existing columns.
        num_columns = self.GetColumnCount()
        for num in range(num_columns - 1):
            width -= self.GetColumnWidth(num)
        # Use CallAfter to try to prevent horizontal scrollbar flicker.
        wx.CallAfter(self.SetColumnWidth, num_columns - 1, max(int(width), 100))

    def OnTreeKeyDown(self, event):
        """Allows us to intercept key presses in the tree control.

        NOTE: Event is actually a TreeEvent not a KeyEvent.
        """
        # Event is a wx.TreeEvent object. Check for delete key.
        keycode = event.GetKeyCode()
        if keycode == wx.WXK_DELETE:
            # DEL key deletes the current selection.
            # wx.EVT_TREE_DELETE_ITEM
            # Delete all items from the HyperTreeList
            # remove the selected item from the added devices list.

            # Re Add all of the devices and then update the hyper tree list
            self.OnTreeDelete(event)
        elif keycode == wx.WXK_WINDOWS_MENU:
            # EVT_TREE_ITEM_MENU places context menu badly. Do it ourselves.
            selections = self.GetSelections()
            if selections:
                # Popup context menu at right edge of item, centered vertically.
                y = selections[0].GetY()
                x, y = self.GetMainWindow().CalcScrolledPosition(0, y)
                rect = self.GetBoundingRect(selections[0], textOnly=True)
                offset = self.GetColumnWidth(0)
                if rect is None:
                    position = (x + offset, y + 9)
                else:
                    position = (x + offset, y + rect.height / 2)
                self.PopupMenu(self.menu, position)
        else:
            # Allow all other key events to pass so that arrow keys work.
            event.Skip(True)

    def OnTreeDelete(self, event):
        '''User pressed DEL or delete menu item. Deletes all selected items.'''
        self.OnDeleteItems(self.GetSelections())

    def OnDeleteItems(self, items):
        '''Delete the given list of items. Can be overridden to prompt
        user for confirmation.'''
        for itemId in items:
            self.DeleteItem(itemId)

    def DeleteItem(self, itemId):
        '''Delete the tree itemId and remove its group, if empty.'''
        # Get GroupId of this item before deleting it.
        groupId = self.GetItemParent(itemId)
        print(f"DEBUG: groupId: {groupId}")
        # Quick check here to see if this item has already been deleted.
        siblings = groupId.GetChildren()
        print(f"DEBUG: siblings: {siblings}")
        if itemId not in siblings:
            # Item already deleted. Workaround if parent deleted before child.
            if self.IsSelected(itemId):
                self.SelectItem(self.GetRootItem())
            return
        #Item Text
        item_ip = itemId.GetText()
        device_found = None
        # Go through each device and check what the device IP is
        for index, device in enumerate(self.parent.added_devices):
            if item_ip == device.IP:
                device_found = index
        if device_found is not None:
            # Remove the device from the list
            self.parent.added_devices.pop(device_found)
        else:
            # They are trying to delete a group.
            wx.MessageBox("Cannot Delete SubModules. Please choose main Device",
                          style=wx.OK | wx.ICON_INFORMATION, parent=self)
            return
        nextItem = None
        if self.IsSelected(itemId):
            # Move selection to next item before deleting.
            nextItem = self.GetNextShown(itemId)
        # Finally, delete the item.
        self.Delete(itemId)
        print("DEBUG: deleted the item!")
        # # If the group the item belonged to is empty, delete it as well.
        if self.GetChildrenCount(groupId) == 0:
            if nextItem or groupId.IsSelected():
                nextItem = self.GetNextShown(groupId)
            self.Delete(groupId)
        # Move selection off deleted item, if it was selected beforehand.
        if nextItem:
            if not self.IsSelected(nextItem):
                # Workaround where SelectItem() actually toggles selection.
                self.SelectItem(nextItem)

    def OnTreeRightClick(self, event: HTL.TreeEvent):
        """On Right click display pop-up menu, if one was defined."""
        if not self.menu:
            return
        self.PopupMenu(self.menu, event.GetPoint())

    def OnFind(self, event):
        """Search all tree items for anything containing search text.

        By default we search the item's data.
            UnmatchedTreeList: data = TemplateCategory or LogLine.
            TemplateTreeList: data = None
            AnomalyTreeList: data = Incident or LogLine
        If that fails we search the base tree text.

        Return wx.NOT_FOUND if the text was not found.
        """
        text = event.GetFindString()
        # Flags can be wx.FR_DOWN, wx.FR_MATCHCASE, wx.FR_WHOLEWORD.
        flags = event.GetFlags()
        item = self.GetSelection()
        item = self.root if item is None or not item.IsOk() else item
        lower = False if flags & wx.FR_MATCHCASE else True
        find_text = text if flags & wx.FR_MATCHCASE else text.lower()
        whole_word = True if flags & wx.FR_WHOLEWORD else False
        down = True if flags & wx.FR_DOWN else False
        while item is not None and item.IsOk():
            # Move to the next item, or next child.
            item = self.GetNext(item) if down is True else self.GetPrev(item)
            if item:
                data = self.GetItemPyData(item)
                if type(data) is self.TemplateCategory:
                    content = data.template
                else:
                    content = self.GetItemText(item)
                content = content.lower() if lower is True else content
                content = content.split() if whole_word is True else content
                if find_text in content:
                    self.EnsureVisible(item)
                    self.SelectItem(item)
                    break
        else:   # No break
            wx.MessageBox('Cannot find "%s"' % text, caption="Find Log",
                          style=wx.OK | wx.ICON_INFORMATION, parent=self)
            item = wx.NOT_FOUND
        return item

    def OnCopy(self, event):
        """Push the contents of the selected lines into the clipboard."""
        item = self.GetSelection()
        if item and item.IsOk():
            data = self.GetItemPyData(item)
            text = None
            if type(data) is self.TemplateCategory:
                text = data.template
            elif type(data) is list:
                text = '\n'.join(logline.original for logline in data)
            if text:
                copy_to_clipboard(text)

    def FindGroup(self, group_name):
        """
            This function is created to find the GroupName in the HyperTreeList.
            It will return the specified HyperTreeListItem.

            Parameter:
                group_name: The text value of the HyperTreeListItem you are looking for
        """
        tree_groups = self.root.GetChildren()
        device_groups = [child.GetText() for child in tree_groups]
        # Try to find the group_name in the root Tree
        try:
            device_index = device_groups.index(group_name)
            # print(f"DEBUG: I have found the Group. Returning it. {device_index}, device_groups:{device_groups}")
        except ValueError:
            print("The group you are looking for is not found in the Tree. Adding Group %s" % (
                group_name))
            # Create the Group if its not in the Tree already.
            return self.AppendItem(self.root, group_name)
        return tree_groups[device_index]

    def SetupColumns(self, device, new_device, device_found=False):
        """A method to populate the columns for the UpgradeListCtrl

        Parameters:
            results: A DeviceInfo object's results. (IP, SysDescr, mac, old firmware)
            new_device: A HyperTreeListItem Object. This will populate a row in the tree.
            device_found: A boolean. If we have a model for the specified device, it will put a
                         checkmark, if it does not it will put an X in the supported row.
        """
        if device_found:
            self.SetItemText(new_device, u'\u2713', column=1)
        else:
            self.SetItemText(new_device, 'X', column=1)
        results = device.results
        self.SetItemText(new_device, str(results.get('card_name',
                                                     device.snmp.get('1.3.6.1.2.1.1.1.0'))),
                                                     column=2)
        self.SetItemText(new_device, str(results.get('serial', '')), column=3)
        self.SetItemText(new_device, str(results.get('mac', '')), column=4)
        self.SetItemText(new_device, str(results.get('alias', '')), column=5)
        current_firmware = '%s.%s.%s' % (results.get('major', ''),
                                         results.get('minor', ''),
                                         results.get('build', ''))
        self.SetItemText(new_device, str(current_firmware), column=6)

    def CheckForDuplicates(self, root_group: HTL.TreeListItem, device: DeviceInfo):
        """
            This function is created to check for duplicate IP's. If there is a duplicate IP,
            it will return True. If no duplicate found, it will return False.

            Parameters:
                root_group: The TreeListItem at which you want to check if the IP is in the list.
                device: A DeviceInfo object.
        """
        tree_groups = root_group.GetChildren()
        device_ips = [child.GetText() for child in tree_groups]
        if device_ips:
            if device.IP in device_ips:
                print(f"DEBUG: Device already in the list. Device.ip: {device.IP}")
                return True
        return False

    def AddDevice(self, device: DeviceInfo):
        """ Adds the device passed to the main panel. If there is a 'modules'
            variable found for the device add them as sub-devices"""
        tree_groups = self.root.GetChildren()
        device_groups = [child.GetText() for child in tree_groups]
        results = device.results
        # Check the group.
        if device.model != None:
            model = device.model.GROUP
            # print(f"DEBUG:AddDevice Model:{model}, ")
            # Found a device. It has its own group.
            if model not in device_groups:
                # print(f"DEBUG:Creating known group Model:{model}, ")
                self.AppendItem(self.root, model)
            # Pull the unkown group HyperTreeListItem
            root_group = self.FindGroup(model)
            # Check if there is already a device with the same IP
            if self.CheckForDuplicates(root_group, device) is True:
                return None
            # Add the device to the group
            self.parent.added_devices.append(device)
            new_device = self.AppendItem(root_group, str(device.IP))
            self.SetupColumns(device, new_device, device_found=True)
            # Check if there are any modules and add them in as children
            if hasattr(device.model,'modules'):
                modules = device.model.modules
            else:
                modules = None
            # Go through device specific object and get the name
            if modules:
                for module in modules:
                    sub_module = self.AppendItem(new_device, str(module.IP))
                    self.SetItemText(sub_module, u'\u2713', column=1)
                    self.SetItemText(sub_module, module.name, column=2)
                    serial = module.results.get('serial', '-')
                    self.SetItemText(sub_module, serial, column=3)
                    mac = module.results.get('mac', '-')
                    self.SetItemText(sub_module, mac, column=4)
                    self.SetItemText(sub_module, module.current_firmware,
                                     column=6)
        else:
            # No device found put them into the unknown group
            if 'UNKNOWN' not in device_groups:
                # print(f"DEBUG:Creating unknown group Model:{model}, ")
                self.AppendItem(self.root, 'UNKNOWN')
            # Pull the unkown group HyperTreeListItem
            root_group = self.FindGroup('UNKNOWN')
            # Check if there is already a device with the same IP
            if self.CheckForDuplicates(root_group, device) is True:
                return None
            # Add the device to the group
            new_device = self.AppendItem(root_group, str(device.IP))
            self.parent.added_devices.append(device)
            self.SetupColumns(device, new_device)
        return

    def export(self, path):
        """Export the contents of the tree to an Excel file."""
        import openpyxl
        import openpyxl.utils
        import openpyxl.styles
        import openpyxl.worksheet.worksheet
        BOTTOM_THIN_BORDER = openpyxl.styles.Border(
            bottom=openpyxl.styles.Side(border_style="thin"))
        # TIER_FILLS = (
        #     (openpyxl.styles.fills.DEFAULT_EMPTY_FILL,) +
        #     tuple(openpyxl.styles.PatternFill(
        #           patternType='solid', fgColor=Color(rgb="%02x%02x%02x" % color[:3]))
        #           for color in TIER_COLORS[1:]))
        RIGHT_ALIGN = openpyxl.styles.alignment.Alignment(horizontal='right')
        HEADER_FONT = openpyxl.styles.fonts.Font(b=True)

        book = openpyxl.Workbook()
        sheet: openpyxl.worksheet.worksheet.Worksheet = book.active
        sheet.title = "DeviceLab Export"
        # Header row
        row = 1
        for col in range(self.GetColumnCount()):
            width = 35 if col == 0 else self.GetColumnWidth(col) / 7
            sheet.cell(row, col + 1).value = self.GetColumnText(col)
            sheet.cell(row, col + 1).font = HEADER_FONT
            sheet.cell(row, col + 1).border = BOTTOM_THIN_BORDER
            letter = openpyxl.utils.get_column_letter(col + 1)
            sheet.column_dimensions[letter].width = width
        row += 1
        sheet.freeze_panes = sheet.cell(row=row, column=1)
        # Dump all devices.
        item, root_cookie = self.GetFirstChild(self.root)
        while item and item.IsOk():
            # Populate this root item's columns.
            for col in range(self.GetColumnCount()):
                sheet.cell(row, col + 1).value = self.GetItemText(item, col)
                sheet.cell(row, col + 1).font = HEADER_FONT
            row += 1
            # Now scan children of this root item.
            child, child_cookie = self.GetFirstChild(item)
            while child and child.IsOk():
                # Populate all columns for this child item.
                for col in range(self.GetColumnCount()):
                    sheet.cell(row, col + 1).value = self.GetItemText(child, col)
                    sheet.cell(row, col + 1).fill = None
                    if col == 0:
                        sheet.cell(row, col + 1).alignment = RIGHT_ALIGN
                row += 1
                child, child_cookie = self.GetNextChild(item, child_cookie)
            # Go to the next root item (if any).
            item, root_cookie = self.GetNextChild(self.root, root_cookie)
        # Finished populating worksheet. Write to file (may raise Exception).
        book.save(open(path, 'wb'))
#end class DeviceTree(HTL.HyperTreeList)


# Main Application Classes --------------------------------------------
class AppPanel(wx.Panel):
    """Panel to show the device."""
    POLL_RATE = 250         # How often (in milliseconds) to poll devices.

    def __init__(self, frame, wxconfig, debug=0):
        wx.Panel.__init__(self, parent=frame)
        self.frame: AppFrame = frame
        self.wxconfig: wx.ConfigBase = wxconfig
        self.find_data = wx.FindReplaceData(wx.FR_DOWN)
        self.find_dialog = None
        self.devices = None
        self.added_devices = []

        # Define UI ------------------------------------------------
        main = wx.BoxSizer(orient=wx.VERTICAL)
        row = wx.BoxSizer(orient=wx.VERTICAL)
        self.gauge = wx.Gauge(self, range=100, size=(-1, 2))
        row.Add(self.gauge, flag=wx.EXPAND)
        main.Add(row, flag=wx.EXPAND | wx.ALL, border=0)
        # The Device Tree. Shows all data and state of all devices.
        device_menu = wx.Menu()
        menu_scan = device_menu.Append(wx.ID_ANY, "&Scan Subnet\tCtrl+O")
        # menu_add = device_menu.Append(wx.ID_ANY, "&Add Device\tCtrl+N")
        menu_find = device_menu.Append(wx.ID_ANY, "&Find...\tCtrl+F")
        menu_copy = device_menu.Append(wx.ID_ANY, "&Copy Line\tCtrl+C")
        device_menu.AppendSeparator()
        menu_remove = device_menu.Append(wx.ID_ANY, "&Remove Device\tDEL")
        device_menu.AppendSeparator()
        menu_export = device_menu.Append(wx.ID_ANY, "&Export to Excel")
        self.Bind(wx.EVT_MENU, self.OnScan, source=menu_scan)
        # self.Bind(wx.EVT_MENU, self.OnAdd, source=menu_add)
        self.Bind(wx.EVT_MENU, self.OnFind, source=menu_find)
        self.Bind(wx.EVT_MENU, self.OnCopy, source=menu_copy)
        self.Bind(wx.EVT_MENU, self.OnDelete, source=menu_remove)
        self.Bind(wx.EVT_MENU, self.OnExport, menu_export)
        self.device_tree = DeviceTree(self, gauge=self.gauge)
        main.Add(self.device_tree, proportion=1, flag=wx.EXPAND)
        # -------------------------------------------------------------
        self.SetSizerAndFit(main)

        # Add keyboard shortcut accelerators for "Find".
        ctrl_f = wx.NewIdRef()
        accel = wx.AcceleratorTable([(wx.ACCEL_CTRL, ord('F'), ctrl_f.GetId()),
                                     ])
        self.SetAcceleratorTable(accel)

        # Create timer to poll the session to do its tasks.
        self.timer = wx.Timer(self)
        self.Bind(wx.EVT_TIMER, self.OnTimer, source=self.timer)
        self.timer.Start(milliseconds=self.POLL_RATE)

        # Bind events
        self.Bind(wx.EVT_MENU, self.OnFind, id=ctrl_f)

    def SetStatusText(self, text, number=0):
        """Delegate calls to parent frame to set statusbar text."""
        self.frame.SetStatusText(text, number)

    def OnTimer(self, event):
        """Poll session so it can maintain any ongoing processes."""
        pass

    def OnScan(self, event=None):
        """Show scan dialog to allow user to scan subnets."""
        # Create and display dialog
        dialog = ScanDialog(self, self.wxconfig)
        dialog.ShowModal()
        dialog.Destroy()

    # def OnAdd(self, event=None):
    #     """Add a single new device by IP address."""
    #     dialog = AddDialog(self, self.wxconfig)
    #     reply = dialog.ShowModal()
    #     dialog.Destroy()

    def OnExport(self, event=None):
        """Export the DeviceTree to an Excel file."""
        # Prompt for a file to export to.
        export_path = self.wxconfig.Read('/ExportPath', os.getcwd())
        wildcard = ("Excel (*.xlsx)|*.xlsx")
        dialog = wx.FileDialog(self, "Choose Excel file to export to",
                               wildcard=wildcard, defaultDir=export_path,
                               style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)
        reply = dialog.ShowModal()
        export_path = dialog.GetPath()
        dialog.Destroy()
        if reply not in (wx.OK, wx.ID_OK) or not export_path:
            return
        # Store the "last export" path and export the DeviceTree.
        self.wxconfig.Write('/ExportPath', os.path.dirname(export_path))
        try:
            self.device_tree.export(export_path)
        except Exception as err:
            # Export Failed. Show error message dialog.
            dialog = wx.RichMessageDialog(self, message="Failed to export\n%s" %
                                          str(err), caption="Export Failed")
            dialog.ShowDetailedText(traceback.format_exc())
            dialog.ShowModal()
            dialog.Destroy()
        else:
            # Export successful. Prompt if user wants to open.
            reply = wx.MessageBox("Export Successful. Open file now?",
                                  caption="Export Successful", style=wx.YES_NO |
                                  wx.CENTER | wx.NO_DEFAULT, parent=self)
            if reply in (wx.YES, wx.ID_YES):
                # Open the exported file right away.
                if sys.platform.startswith('win'):
                    os.startfile(export_path)
                elif sys.platform.startswith('darwin'):
                    subprocess.call(('open', export_path))
                else:
                    subprocess.call(('xdg-open', export_path))

    def OnExit(self, event):
        """Called with event when user tries to close main frame.

        Can veto application exit with event.Skip(skip=False).
        """
        if self.timer.IsRunning():
            self.timer.Stop()

    def OnFind(self, event):
        """Show FindDialog and search logs for given text."""
        if not self.find_dialog:
            self.find_dialog = wx.FindReplaceDialog(self, self.find_data,
                                                    title="Find Log Text")
            self.find_dialog.Bind(wx.EVT_FIND, self.OnFindText)
            self.find_dialog.Bind(wx.EVT_FIND_NEXT, self.OnFindText)
            self.find_dialog.Bind(wx.EVT_FIND_CLOSE, self.OnFindClose)
        self.find_dialog.Show()
        self.find_dialog.SetFocus()

    def OnFindText(self, event):
        """Handle a FindEvent by forwarding it to the active top book page."""
        page = self.top_book.GetCurrentPage()
        if page in (self.log_list, self.templates_tree, self.unmatched_tree,
                    self.anomaly_tree):
            print("OnFindText %s" % page.__class__.__name__)
            index = page.OnFind(event)
            # If nothing found, set focus back to the find dialog.
            if index == wx.NOT_FOUND and self.find_dialog:
                self.find_dialog.SetFocus()

    def OnFindClose(self, event):
        self.find_dialog.Destroy()
        self.find_dialog = None

    def OnCopy(self, event):
        """Copy the currently selected device in the DeviceTree to Clipboard."""
        pass

    def OnDelete(self, event):
        """Remove the currently selected device in the DeviceTree."""
        pass

    def OnUpgrade(self, event: wx.CommandEvent):
        """ Brings up Dialog which contains the upgrade procedure"""
        # Set the devices added to the DeviceTree as the device
        self.devices = self.added_devices
        self.supported_devices = []
        dialog = UpgradeDialog(self, self.wxconfig)
        dialog.ShowModal()

    def OnLicense(self, event: wx.CommandEvent):
        """
            Brings up Dialog which contains the upgrade procedure
        """
        dialog = LicenseDialog(self, self.wxconfig)
        dialog.ShowModal()

    def OnClose(self, event=None, warn_only=False):
        """Close current session, if one exists.

        'warn_only' if set to True will only warn the user if they
                    unsaved changes and allow them to save. Won't
                    actually close the session.
        """
        if self.session:
            if self.session.changed is True:
                reply = wx.MessageBox("Save session before closing?",
                                      caption="Unsaved Changes",
                                      style=wx.YES_NO | wx.ICON_EXCLAMATION)
                if reply in (wx.YES, wx.OK, wx.ID_YES, wx.ID_OK):
                    self.OnSave()
        self.UpdateUI()
#end class SessionPanel(wx.Panel)

class AppFrame(wx.Frame):
    """Main application frame (window)"""
    def __init__(self, debug=0):
        """Initialize our main application frame."""
        # Call the original constructor to do its job
        UNIT = max(int(wx.SystemSettings.GetMetric(wx.SYS_ICON_X) * 1.125), 36)
        TITLE = "%s v%s" % (PRODUCTNAME, VERSION)
        wx.Frame.__init__(self, parent=None, title=TITLE, size=(1366, 768))

        # Setup application icon bundle
        self.icons = wx.IconBundle()
        self.icons.AddIcon(images.DeviceLab16.GetIcon())
        self.icons.AddIcon(images.DeviceLab32.GetIcon())
        self.icons.AddIcon(images.DeviceLab64.GetIcon())
        self.icons.AddIcon(images.DeviceLab128.GetIcon())
        self.SetIcons(self.icons)

        # Load our wxconfig settings (\HKCU\Software\vendorName\appName).
        self.wxconfig = wx.Config()

        # Create a status bar (displays help and memory usage).
        self.CreateStatusBar(number=3, style=wx.STB_DEFAULT_STYLE)
        self.SetStatusWidths([-1, UNIT * 5, UNIT * 6])

        # Create our menu bar.
        menu_bar = wx.MenuBar()
        file_menu = wx.Menu()
        file_scan = file_menu.Append(wx.MenuItem(
            file_menu, id=wx.ID_ANY, text="&Scan Subnet\tCtrl+O",
            helpString="Scan a subnet for devices"))
        file_add = file_menu.Append(wx.MenuItem(
            file_menu, id=wx.ID_ANY, text="&Add Device\tCtrl+N",
            helpString="Add a single device by IP address"))
        file_export = file_menu.Append(wx.MenuItem(
            file_menu, id=wx.ID_ANY, text="E&xport\tCtrl+S",
            helpString="Export device list to Excel file"))
        file_menu.AppendSeparator()
        self.file_exit = file_menu.Append(wx.MenuItem(
            file_menu, id=wx.ID_ANY, text="E&xit\tAlt+F4",
            helpString="Exit program"))
        menu_bar.Append(file_menu, "&File")
        tools_menu = wx.Menu()
        tools_upgrade = tools_menu.Append(wx.MenuItem(
            tools_menu, wx.ID_ANY, text="Upgrade...", kind=wx.ITEM_NORMAL,
            helpString="Upgrade devices"))
        tools_licensing = tools_menu.Append(wx.MenuItem(
            tools_menu, wx.ID_ANY, text="Licensing...", kind=wx.ITEM_NORMAL,
            helpString="License devices"))
        menu_bar.Append(tools_menu, "&Tools")
        help_menu = wx.Menu()
        help_about = help_menu.Append(wx.MenuItem(
            help_menu, wx.ID_ANY, text="&About\tF1", kind=wx.ITEM_NORMAL,
            helpString="Program Information"))
        menu_bar.Append(help_menu, "&Help")
        self.SetMenuBar(menu_bar)

        # Initialize UI - AppPanel
        sizer = wx.BoxSizer(orient=wx.VERTICAL)
        self.app_panel = AppPanel(self, self.wxconfig, debug=debug)
        sizer.Add(self.app_panel, proportion=1, flag=wx.EXPAND)
        self.SetSizer(sizer)

        # Create MenuItem list
        # Generate toolbar
        self.CreateToolBar(style=wx.TB_HORIZONTAL | wx.TB_TEXT | wx.TB_FLAT)
        tb = self.GetToolBar()
        # tb.SetMargins((50,100))
        size = wx.ArtProvider.GetNativeSizeHint(wx.ART_TOOLBAR)
        # Scan, Export, Upgrade, license,
        CLIENT = wx.ART_TOOLBAR
        tb.AddSeparator()
        bitmap = wx.ArtProvider.GetBitmap(wx.ART_FIND, CLIENT, size)
        toolScan = tb.AddTool(wx.ID_ANY, "Scan", bitmap, "Scan Subnet")
        tb.SetToolLongHelp(toolScan.GetId(), "Scan a Subnet for devices")
        self.Bind(wx.EVT_MENU, self.app_panel.OnScan, source=toolScan)
        tb.AddSeparator()
        # Export
        bitmap = wx.ArtProvider.GetBitmap(wx.ART_UNDO, CLIENT, size)
        toolExport = tb.AddTool(wx.ID_ANY, "Export", bitmap, "Export Table")
        tb.SetToolLongHelp(toolExport.GetId(), "Export table into excel format")
        self.Bind(wx.EVT_MENU, self.OnAbout, source=toolExport)
        tb.AddSeparator()
        # Upgrade
        bitmap = wx.ArtProvider.GetBitmap(wx.ART_GO_UP, CLIENT, size)
        toolUpgrade = tb.AddTool(wx.ID_ANY, "Upgrade", bitmap, "Upgrade Devices")
        tb.SetToolLongHelp(toolUpgrade.GetId(), ("Brings up a Dialog to upgrade "
                           "devices based on upgrade file chosen"))
        self.Bind(wx.EVT_MENU, self.app_panel.OnUpgrade, source=toolUpgrade)
        tb.AddSeparator()
        # license
        bitmap = wx.ArtProvider.GetBitmap(wx.ART_HELP_PAGE, CLIENT, size)
        toolLicense = tb.AddTool(wx.ID_ANY, "License", bitmap, "Add License")
        tb.SetToolLongHelp(toolLicense.GetId(), ("Brings up a Dialog to add a license to a device "))
        self.Bind(wx.EVT_MENU, self.app_panel.OnLicense, source=toolLicense)
        tb.AddSeparator()
        tb.Realize()

        # Bind events
        self.Bind(wx.EVT_CLOSE, self.OnClose)
        self.Bind(wx.EVT_MENU, self.app_panel.OnScan, source=file_scan)
        # self.Bind(wx.EVT_MENU, self.app_panel.OnAdd, source=file_add)
        self.Bind(wx.EVT_MENU, self.app_panel.OnExport, source=file_export)
        self.Bind(wx.EVT_MENU, self.app_panel.OnUpgrade, source=tools_upgrade)
        self.Bind(wx.EVT_MENU, self.app_panel.OnLicense, source=tools_licensing)
        self.Bind(wx.EVT_MENU, self.OnExit, source=self.file_exit)
        self.Bind(wx.EVT_MENU, self.OnAbout, source=help_about)

        # Show our main window.
        self.Show()

    def OnExit(self, event=None):
        """Exit the program. Frame.Close() generates a EVT_CLOSE event."""
        self.Close()

    def OnClose(self, event: wx.CloseEvent):
        """User wants to close the application. Forward to app_panel."""
        # Skip event by default so it propagates, closing the application.
        event.Skip()
        # Send event to AppPanel/BookPages. They can veto by clearing skip flag.
        self.app_panel.OnExit(event)
        # If we are exiting, stop the timer and loader processes.
        if event.GetSkipped():
            # Perform any cleanup here.
            pass

    def OnAbout(self, event=None):
        """Create and show an AboutDialogInfo() dialog with program info."""
        info = wx.adv.AboutDialogInfo()
        info.SetName(VERSIONINFO['fileDescription'])
        info.SetVersion(str(VERSION))
        info.SetDescription(
            "Evertz employee use only. Do not distribute.\n" +
            "Python version %s.%s.%s (%s %s)\n" % tuple(sys.version_info) +
            "Powered by wxPython %s\n" % (wx.version()) +
            "Running on %s\n\n" % (wx.GetOsDescription()) +
            "Process ID = %s\n" % (os.getpid()))
        info.SetCopyright(VERSIONINFO['legalCopyright'])
        info.SetWebSite("www.evertz.com", VERSIONINFO['companyName'])
        info.SetIcon(images.EvertzGold256.GetIcon())
        # Show the generic About Dialog
        wx.adv.AboutBox(info)
#end class MainFrame(wx.Frame)

class UpgradeDialog(wx.Dialog):
    """Dialog to provide upgrade ability on devices based on the file the user passes.

    self.upgrade_firmware_path is the filepath to the file on the users pc
    self.upgrade_file: is the actual file the user is using to look for the possible upgrade devices
    self.firmware_info: a touple (firmware_device, firmware_version) of the device
    self.num_selected : Number of devices selected for upgrade.

    .ListCtrl for each device discovered.
    The user data for each item contains the DeviceInfo object.
    """
    def __init__(self, parent: wx.Panel, wxconfig: wx.ConfigBase):
        wx.Dialog.__init__(self, parent, title="Scan Subnets for Devices",
                           style=wx.DEFAULT_DIALOG_STYLE | wx.RESIZE_BORDER,
                           size=(1000, 2000))
        # self.SetSize((1000,1000))
        self.SetMinSize((1200, 400))
        # self.Refresh()
        self.parent: wx.Panel = parent
        self.wxconfig: wx.ConfigBase = wxconfig
        self.supported_devices: typing.List[DeviceInfo] = []
        self.num_selected = 0
        # If the device is supported it provides some information on the firmware.
        self.firmware_info = None
        title_font = wx.Font(18, wx.DECORATIVE, wx.ITALIC, wx.NORMAL)
        # Initialize UI -----------------------------------------------
        UNIT = max(int(wx.SystemSettings.GetMetric(wx.SYS_ICON_X) * 1.125), 36)
        ACVR = wx.ALIGN_CENTER_VERTICAL | wx.ALL
        main = wx.BoxSizer(orient=wx.VERTICAL)

        # Create a Spacer between Top of the dialog and Title
        main.Add(wx.StaticBox(self, size=(1, 10)))
        # Title Row
        # upgrade_title_row = wx.BoxSizer(orient=wx.HORIZONTAL)
        upgrade_main_label = wx.StaticText(self, label="Device Upgrade",
                                           style=wx.ALIGN_CENTER)
        upgrade_main_label.SetFont(title_font)
        main.Add(upgrade_main_label, flag=wx.EXPAND)
        # Create a Spacer
        main.Add(wx.StaticBox(self, size=(1, 20)))
        # Choose Upgrade File Row
        choose_file_row = wx.BoxSizer(orient=wx.HORIZONTAL)
        choose_file_row.AddStretchSpacer()
        upgrade_choose_file_label = wx.StaticText(self, label="Choose Upgrade File: ")
        choose_file_row.Add(upgrade_choose_file_label, flag=ACVR)
        self.choose_file_button = wx.Button(self, label="&Choose File")
        self.choose_file_button.SetToolTip("Choose a File")
        choose_file_row.Add(self.choose_file_button)
        choose_file_row.AddStretchSpacer()
        main.Add(choose_file_row, border=UNIT * 6, flag=wx.EXPAND)
        # Add row specifying the file that they have chosen
        row = wx.BoxSizer(orient=wx.HORIZONTAL)
        row.AddStretchSpacer()
        self.chosen_file_label = wx.StaticText(self, label="File Chosen: ", )
        self.chosen_file_label.Hide()
        row.Add(self.chosen_file_label, flag=ACVR)
        # Will be populated when the user chooses a file. Contains the name of the file we update
        self.upgrade_chosen_file_label = wx.StaticText(self, label="")
        row.Add(self.upgrade_chosen_file_label)
        row.AddStretchSpacer()
        main.Add(row, border=UNIT * 6, flag=wx.EXPAND)
        # Create a Spacer between Choose Upgrade and Supported Device CtrList
        main.Add(wx.StaticBox(self, size=(1, 20)))
        # Add title for upgrade list
        row = wx.BoxSizer(orient=wx.HORIZONTAL)
        row.Add(wx.StaticBox(self, size=(100, 1)))
        upgrade_list_label = wx.StaticText(self, label="Compatable Devices: ")
        row.Add(upgrade_list_label)
        main.Add(row, flag=wx.EXPAND)
        # Add Upgrade list
        # self.upgrade_list = wx.BoxSizer(orient=wx.HORIZONTAL)
        # Create the Supported Devices List.
        row = wx.BoxSizer(orient=wx.HORIZONTAL)
        row.AddStretchSpacer()
        self.upgrade_list = UpgradeList(self, self.supported_devices)
        row.Add(self.upgrade_list)
        row.AddStretchSpacer()
        main.Add(row, flag=wx.EXPAND)

        # Create a Spacer between Upgrade CtrList and Start upgrade button
        main.Add(wx.StaticBox(self, size=(1, 20)))
        # Add the Start Upgrade Button
        row = wx.BoxSizer(orient=wx.HORIZONTAL)
        self.start_upgrade_button = wx.Button(self, label="&Start Upgrade")
        self.start_upgrade_button.SetToolTip("Start the Upgrade")
        row.AddStretchSpacer()
        row.Add(self.start_upgrade_button)
        # main.Add(row, border=UNIT // 6)
        main.Add(row, border=UNIT * 6, flag=wx.EXPAND)
        row.AddStretchSpacer()
        # Create a Spacer between Start upgrade button and bottom of dialog
        main.Add(wx.StaticBox(self, size=(1, 20)))
        self.SetSizer(main)
        self.Fit()

        self.start_upgrade_button.Bind(wx.EVT_BUTTON, self.OnStartUpgrade)
        self.choose_file_button.Bind(wx.EVT_BUTTON, self.OnUpload)
        #self.SetAffirmativeId(wx.ID_OK)

    def refresh_supported_devices(self):
        """
            This method is created to refresh the Upgrade List Control.
        """
        self.upgrade_list.SetItemCount(len(self.supported_devices))
        # Sort devices by numerical IP address
        self.supported_devices.sort(key=lambda item: item.IP32)
        self.upgrade_list.Update()

    def ValidateUpgradeFile(self, model):
        """
        returns a touple (expected_device, firmware_version)
            expected_device: The expected device the user wants.
            firmware_version: The firmware version you are upgrading to.
        """
        supported = False
        try:
            supported = model.validate_upgrade_firmware(self.firmware_path)
            print(f"DEBUG: firmware validation passed: {supported}"
                  f"path: {self.firmware_path} ")
            return supported
        except Exception as error:
            # traceback.print_exc()
            print("There was a problem validating the " +
                  f"firmware on device {model.IP}, card: {model.name} " +
                  f"{self.firmware_path}, error: {error}")
        return supported

    def OnUpload(self, event):
        """ Provides a FileDialog so the user chooses a file.
            This function validates looks through our devices list and will
            go through each device's validate_upgrade_firmware method which
            validates if this file fits the description for the upgrade

            This function populates:
                self.supported_devices: with devices that are validated/supported
                self.firmware_path: The path of the selected file.
                self.firmware_file: This variable will hold the whole filepath

            This function changes UI:
                chosen_file_label: Displays 'File Chosen: '
                upgrade_chosen_file_label: Displays only the filename


        """
        self.supported_devices = []
        self.chosen_file_label.Show(False)
        self.upgrade_chosen_file_label.Show(False)
        self.upgrade_list.DeleteAllItems()
        self.upgrade_list.Update()
        # otherwise ask the user what new file to open
        with wx.FileDialog(self, "Open Upgrade file", wildcard="Upgrade files (*)|*",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return     # the user changed their mind

            # Proceed loading the file chosen by the user
            self.firmware_path = fileDialog.GetPath()
            try:
                self.firmware_file = open(self.firmware_path, 'rb')
            except IOError:
                wx.LogError("Cannot open file '%s'." % self.firmware_path)
        # Check if the user scanned already
        devices = self.parent.added_devices
        if devices:
            # Show the file Chosen Label
            self.chosen_file_label.Show(True)
            self.upgrade_chosen_file_label.Show(True)
            self.upgrade_chosen_file_label.SetLabelText(str(self.firmware_path.split('\\')[-1]))
            self.Layout()
            # Check if the devices selected are supported
            len_devices = len(devices)
            # Start the progress bar
            percent = 2
            validation_progress = wx.ProgressDialog("File Validation",
                                                    "Searching Devices",
                                                    parent=None,
                                                    style=wx.PD_SMOOTH |
                                                    wx.PD_AUTO_HIDE)
            for dev_num, device in enumerate(devices):
                percent = round(dev_num / len_devices * 100)
                if percent == 0:
                    percent = 5
                validation_progress.Update(percent, 'Starting Validation for '
                                           f'Device: {dev_num}/{len_devices}')
                if device.model:
                    # Check if device has modules (sub-devices)
                    validation_progress.Update(percent, ('Validating Sub Modules'
                                                         f'Device: {dev_num}'
                                                         f'/{len_devices}'))
                    try:
                        # Check if the modules are supported by the file
                        if hasattr(device.model, 'modules'):
                            for module in device.model.modules:
                                print(f"DEBUG: module: {module.name}")
                                supported = self.ValidateUpgradeFile(module)
                                if supported:
                                    print(f"DEBUG: inside OnUpload supported:{supported}")
                                    module.selected = True
                                    self.supported_devices.append(module)
                                    self.refresh_supported_devices()
                                    print(f"DEBUG: module: {module} supported devices: {self.supported_devices}")
                    except AttributeError as error:
                        # If device has no modules continue
                        print(f"Info: No modules found. Error: {error}")
                    # Check if the device itself can be upgraded using the file
                    supported = self.ValidateUpgradeFile(device.model)
                    if not supported:
                        continue
                    # If its supported add it to the supported devices
                    device.model.selected = True
                    self.supported_devices.append(device.model)
                    # Update the virtual list control with the number of devices.
                    self.refresh_supported_devices()
                    # firmware_info ex (firmware_device, firmware_version)
                    self.firmware_info = supported
                    wx.SafeYield()
            # Check if the firmware file is in the correct format
        else:
            self.ClearAll()
            wx.LogError("There are no devices found. Please scan a subnet")
        if validation_progress.GetValue() != 100:
            validation_progress.Update(100, 'Finished Validation')
        validation_progress.Destroy()
        # if self.supported_devices is not populated
        if not self.supported_devices:
            self.ClearAll()
            wx.LogError(("There are no devices found which match the firmware " +
                         "you have chosen firmware passed. " +
                         f"firmware: {self.firmware_path}"))
        # Refresh the windows
        self.Refresh()

    def GetSelectedDevices(self):
        """
            Sets the number of devices that are selected to be upgraded

            Updates self.num_selected
        """
        num_selected = 0
        # Count how many devices they want to upgrade.
        for device in self.supported_devices:
            if device.selected:
                num_selected += 1
        return num_selected

    def OnStartUpgrade(self, event):
        """ This method will instantiate the upgrade process on the cards in self.supported_devices.
        """
        devices_str = ""
        firmware_version = ""
        expected_device = ""
        # List holding completed devices
        # completed_upgrades = []
        # Get the number of devices
        self.num_selected = self.GetSelectedDevices()
        if self.num_selected < 1:
            wx.MessageDialog(self, "No devices chosen to upgrade. " +
                             f"Number of selected devices: {self.num_selected}",
                             caption="No Devices Found!",
                             style=wx.OK | wx.ICON_WARNING).ShowModal()
            return None
        # Get the number of devices that we are upgrading
        devices_to_upgrade = []
        # Go through each supported device
        for device in self.supported_devices:
            # Only upgrade if the device has a selected checkbox
            if device.selected is False:
                continue
            # Go into the device model and attempt to instantiate the upgrade method.
            if 'module' in device.GROUP.lower():
                supported = self.ValidateUpgradeFile(device)
                print(f"DEBUG: Module supported:{supported}")
            else:
                # Add it into one variable so it can be used later.
                supported = self.ValidateUpgradeFile(device)
                print(f"DEBUG: Device supported:{supported}")
            print(f"DEBUG: Supported:{supported}")
            firmware_version = supported[0]
            expected_device = supported[1]
            if not supported:
                # Should never reach this statement
                wx.MessageDialog(self, ("The File which you chose to upgrade " +
                                 f"device: {self.device.model.name}, file: " +
                                 f"{self.firmware_path} is not supported"),
                                 style=wx.ICON_ERROR).ShowModal()
                continue
            # if there is a device that is
            devices_to_upgrade.append(device)
            # Create a dialog with the upgraded devices.
            print(f"DEBUG: outside devices_to_upgrade:{devices_to_upgrade}")
        # Display the information to the user one last time
        result = DisplayUpgradeInfoDialog(self, self.wxconfig, devices_to_upgrade).ShowModal()
        print(f"DEBUG: outside user_result:{result}")

        # If the user clicks yes.
        if int(result) == 5100:
            polling = []
            percent = 5
            upgrade_progress = wx.ProgressDialog("Starting Upgrades",
                                                 "Validating Version",
                                                 parent=None,
                                                 style=wx.PD_SMOOTH |
                                                 wx.PD_AUTO_HIDE)
            upgrade_progress.Update(percent, ("Starting device upgrades "
                                    f"1/{len(devices_to_upgrade)}"))
            for device in devices_to_upgrade:
                # Check if the device is using the Model object or DeviceInfo obj
                # if type(device) is DeviceInfo:
                #     device = device.model
                # Only upgrade if the device has a selected checkbox
                if device.selected is False:
                    continue
                # 1) Start the upgrade thread
                thread = UpgradeThread(device, self.firmware_path, self.firmware_file,
                                       devices_to_upgrade)
                thread.start()
                polling.append(device)

            total_devices = len(polling)
            percent = 10
            upgrade_progress.Update(percent, (f"{total_devices} Device upgrades "
                                    "started. Waiting for devices to "
                                    f"finish upgrading. {total_devices}"
                                    f"/{len(devices_to_upgrade)}"))

            # 2) Poll if the device has finished upgrading
            while polling:
                # 2) After each upgrade is completed update the progress bar
                # Start a timout of 8 minutes
                # start_time = time.time()
                # end_time = start_time + 480
                # Check if the upgrade finished
                for device in devices_to_upgrade:
                    time.sleep(10)
                    print(f"DEBUG: Wait 10 seconds DEVICE UPGRADE FINISHED FLAG: {device.flags['upgradeFinished']}, error:{device.flags['upgradeError']}")
                    if device.flags['upgradeFinished'] is True or device.flags['upgradeError']:
                        # Update the progress bar
                        percent += int((len(polling) / total_devices) * 100)
                        if percent > 100:
                            percent = 98
                            upgrade_progress.Update(percent, (f"{total_devices} "
                                                "Waiting for devices to "
                                                "finish upgrading. "
                                                f"{len(polling)}/"
                                                f"{len(devices_to_upgrade)}"))
                        if device in polling:
                            # Remove the device from list
                            polling.remove(device)
                        else:
                            print(f"ERROR: Device was not found in polling. {device.IP}")
            # Check if all devices passed.
            upgrade_progress.Update(100)
            upgrade_progress.Destroy()
            DisplayUpgradeInfoDialog(self, self.wxconfig, devices_to_upgrade,
                                     first_display=False).ShowModal()
        else:
            # There was a problem starting the upgrade.
            print("User decided against starting the upgrade.")
            return None

    def ClearAll(self):
        """This method will remove everything from the UpgradeList Control"""
        self.supported_devices = []
        self.upgrade_list.DeleteAllItems()
        self.upgrade_list.Update()
#end class Upgrade Dialog

class UpgradeList(wx.ListCtrl):
    """List control to show discovered device information.

    Expects Parent to have values populated for:
        self.parent.firmware_info
        self.parent.supported_devices

    self.devices = List of DeviceInfo class to show.
    NOTE: device needs to have device.selected to work.

    If OnUpload (OnOpen) has been run these will be populated.
        self.chosen_filename: The name of the chosen file
        self.image_file: Contains the file object
    """
    COLUMNS = (
            ("Checkmark", wx.LIST_FORMAT_CENTER, 1.5),
            ("IP", wx.LIST_FORMAT_LEFT, 3.6),
            ("Description", wx.LIST_FORMAT_CENTRE, 4.3),
            ("Serial", wx.LIST_FORMAT_CENTRE, 5.0),
            ("MAC", wx.LIST_FORMAT_CENTRE, 5.0),
            ("Current Firmware", wx.LIST_FORMAT_CENTRE, 4.0),
            ("Chosen Firmware", wx.LIST_FORMAT_CENTRE, 7.0))

    def __init__(self, parent, devices: dict, menu=None, style=0):
        # Call the original constructor to do its job. Force style
        style |= wx.LC_REPORT | wx.LC_VIRTUAL
        wx.ListCtrl.__init__(self, parent, style=style)
        # Assign our parent book and figure out our page number.
        self.parent: wx.Window = parent             # Parent window
        self.menu: wx.Menu = menu                   # Popup context menu.
        # This is all devices in our list control.
        self.supported_devices: typing.Dict[int, DeviceInfo] = devices
        # Add standard columns
        UNIT = max(int(wx.SystemSettings.GetMetric(wx.SYS_ICON_X) * 1.125), 36)
        for title, alignment, width in self.COLUMNS:
            self.AppendColumn(title, format=alignment, width=int(width * UNIT))
        # self.Bind(wx.EVT_RIGHT_DOWN, self.OnRightDown)
        # Create Checkbox images
        self.il = wx.ImageList(16, 16)
        self.check_image = self.il.Add(self.CreateBitmap(wx.CONTROL_CHECKED, (16, 16)))
        self.uncheck_image = self.il.Add(self.CreateBitmap(0, (16, 16)))
        self.SetImageList(self.il, wx.IMAGE_LIST_SMALL)

        # Bind events
        self.Bind(wx.EVT_LEFT_DOWN, self.OnLeftDown)
        #self.Bind(wx.EVT_SIZE, self.OnSize)
        self.Bind(wx.EVT_RIGHT_DOWN, self.OnRightDown)

    def CreateBitmap(self, flag=0, size=(16, 16)):
        """Create a bitmap of the platforms native checkbox. The flag
        is used to determine the checkboxes state (see wx.CONTROL_*)
        """
        bmp = wx.Bitmap(*size)
        dc = wx.MemoryDC(bmp)
        dc.SetBackground(wx.WHITE_BRUSH)
        dc.Clear()
        wx.RendererNative.Get().DrawCheckBox(self, dc,
                                             (0, 0, size[0], size[1]), flag)
        dc.SelectObject(wx.NullBitmap)
        return bmp

    def OnGetItemText(self, item_num, column):
        """Return text for the given item number (row) and column number.
           NOTE: With this method we are passed the DeviceInfo object.
        """
        if item_num >= 0 and item_num < len(self.parent.supported_devices):
            model_obj = self.parent.supported_devices[item_num]
            if column == 0:
                # Create a sizer and place the checkbox inside
                # sizer = wx.BoxSizer(orient=wx.VERTICAL)
                # sizer.Add(wx.CheckBox(self, wx.ID_ANY, ("")))
                # self.InsertImageStringItem()
                return ""
            if column == 1:         # Device IP Address
                return model_obj.IP
            if column == 2:         # Card Name/Description of the card
                # If the device is a submodule, get that models name
                # try:
                #     if "module" in model_obj.GROUP.lower():
                #         return model_obj.name
                # except AttributeError:
                #     return model_obj.model.name
                return model_obj.name
            if column == 3:         # Serial Number of the card
                return model_obj.results.get("serial", '-')
            if column == 4:         # Mac Address of the card
                return model_obj.results.get("mac", '-')
            if column == 5:         # Current Firmware on the card
                # current_firmware = '%s.%s.%s' % (
                #                         model_obj.results.get('major', ''),
                #                         model_obj.results.get('minor', ''),
                #                         model_obj.results.get('build', ''))
                return model_obj.current_firmware
            if column == 6:         # Chosen_Firmware
                try:
                    chosen_firmware = self.parent.firmware_info
                except AttributeError as error:
                    print("Warning: No Firmware found There was a problem. "
                          f"error: {error}")
                    chosen_firmware = '-'
                if chosen_firmware:
                    return chosen_firmware[1]
                else:
                    # Get the path (at this point it should be set)
                    return self.parent.firmware_path.split('\\')[-1]
                    # return "-"
        # Unhandled row/column. Return blank string.
        return ""

    def OnGetItemImage(self, item):
        """Return the item checkbox image for the given integer item index."""
        devices = self.parent.supported_devices
        if item >= 0 and item < len(devices):
            device = devices[item]
            try:
                return self.check_image if device.selected else self.uncheck_image
            except Exception:
                print("Selected was not set up")
        # Outside of list! Use no image.
        return -1

    def OnLeftDown(self, event):
        """Snoop left mouse clicks to toggle checkbox icon of items."""
        devices = self.parent.supported_devices
        # Always skip event so it gets processed normally.
        event.Skip()
        # Find out which item, and where, was clicked.
        index, flags = self.HitTest(event.GetPosition())
        if (index >= 0 and index < len(devices) and
                flags == wx.LIST_HITTEST_ONITEMICON):
            # Set the selected state as the opposite of origial state
            device = devices[index]
            device.selected = not device.selected
            self.RefreshItem(index)

    def OnRightDown(self, event):
        """ On Right click, open context menu that can modify selection """
        menu = wx.Menu()
        # Select All
        menu_item = menu.Append(-1, "Select All")
        self.Bind(wx.EVT_MENU, lambda event: self.OnSelectAll(event), menu_item)
        # Unselect All
        menu_item = menu.Append(-1, "Unselect All")
        self.Bind(wx.EVT_MENU, lambda event: self.OnUnselectAll(event), menu_item)
        self.PopupMenu(menu, event.GetPosition())

    def OnUnselectAll(self, event):
        # Get the supported devices and uncheck the fields
        for device in self.parent.supported_devices:
            # un-select the devices
            device.selected = False
        # Update the window
        self.Refresh()

    def OnSelectAll(self, event):
        # Get the supported devices and uncheck the fields
        for device in self.parent.supported_devices:
            # un-select the devices
            device.selected = True
        # Update the window
        self.Refresh()

#end class UpgradeList

class LicenseDialog(wx.Dialog):
    """Dialog to provide Licencing ability on devices based on the file the user passes.

    self.upgrade_license_path is the filepath to the file on the users pc
    self.upgrade_file: is the actual file the user is using to look for the possible upgrade devices
    self.license_info: a touple (size, date) of the license file
    self.num_selected : Number of devices selected for upload.

    .ListCtrl for each device discovered.
    The user data for each item contains the DeviceInfo object.
    """
    def __init__(self, parent: wx.Panel, wxconfig: wx.ConfigBase):
        wx.Dialog.__init__(self, parent, title="Scan Subnets for Devices",
                           style=wx.DEFAULT_DIALOG_STYLE | wx.RESIZE_BORDER,
                           size=(1000, 2000))
        # self.SetSize((1000,1000))
        self.SetMinSize((1200, 400))
        # self.Refresh()
        self.parent: wx.Panel = parent
        self.wxconfig: wx.ConfigBase = wxconfig
        self.supported_devices: typing.List[DeviceInfo] = []
        self.num_selected = 0
        # If the device is supported it provides some information on the license.
        self.license_info = None
        title_font = wx.Font(18, wx.DECORATIVE, wx.ITALIC, wx.NORMAL)
        # Initialize UI -----------------------------------------------
        UNIT = max(int(wx.SystemSettings.GetMetric(wx.SYS_ICON_X) * 1.125), 36)
        ACVR = wx.ALIGN_CENTER_VERTICAL | wx.ALL
        main = wx.BoxSizer(orient=wx.VERTICAL)

        #### TESTING BUTTON ####
        # test_button = wx.Button(self, label="test")
        # test_button.Bind(wx.EVT_BUTTON, self.OnTest)
        # main.Add(test_button)
        #### TESTING BUTTON ####

        # Create a Spacer between Top of the dialog and Title
        main.Add(wx.StaticBox(self, size=(1, 10)))
        # Title Row
        # license_title_row = wx.BoxSizer(orient=wx.HORIZONTAL)
        license_main_label = wx.StaticText(self, label="Device License",
                                           style=wx.ALIGN_CENTER)
        license_main_label.SetFont(title_font)
        main.Add(license_main_label, flag=wx.EXPAND)
        # Create a Spacer
        main.Add(wx.StaticBox(self, size=(1, 20)))
        # Choose license File Row
        choose_file_row = wx.BoxSizer(orient=wx.HORIZONTAL)
        choose_file_row.AddStretchSpacer()
        license_choose_file_label = wx.StaticText(self, label="Choose License File: ")
        choose_file_row.Add(license_choose_file_label, flag=ACVR)
        self.choose_file_button = wx.Button(self, label="&Choose License File")
        self.choose_file_button.SetToolTip("Choose a File")
        choose_file_row.Add(self.choose_file_button)
        choose_file_row.AddStretchSpacer()
        main.Add(choose_file_row, border=UNIT * 6, flag=wx.EXPAND)
        # Add row specifying the file that they have chosen
        row = wx.BoxSizer(orient=wx.HORIZONTAL)
        row.AddStretchSpacer()
        self.chosen_file_label = wx.StaticText(self, label="File Chosen: ", )
        self.chosen_file_label.Hide()
        row.Add(self.chosen_file_label, flag=ACVR)
        # Will be populated when the user chooses a file. Contains the name of the file we update
        self.license_chosen_file_label = wx.StaticText(self, label="")
        row.Add(self.license_chosen_file_label)
        row.AddStretchSpacer()
        main.Add(row, border=UNIT * 6, flag=wx.EXPAND)
        # Create a Spacer between Choose license and Supported Device CtrList
        main.Add(wx.StaticBox(self, size=(1, 20)))
        # Add title for license list
        row = wx.BoxSizer(orient=wx.HORIZONTAL)
        row.Add(wx.StaticBox(self, size=(100, 1)))
        license_list_label = wx.StaticText(self, label="Compatable Devices: ")
        row.Add(license_list_label)
        main.Add(row, flag=wx.EXPAND)
        # Add license list
        # self.license_list = wx.BoxSizer(orient=wx.HORIZONTAL)
        # Create the Supported Devices List.
        row = wx.BoxSizer(orient=wx.HORIZONTAL)
        row.AddStretchSpacer()
        self.license_list = UpgradeList(self, self.supported_devices)
        row.Add(self.license_list)
        row.AddStretchSpacer()
        main.Add(row, flag=wx.EXPAND)

        # Create a Spacer between license CtrList and Start license button
        main.Add(wx.StaticBox(self, size=(1, 20)))
        # Add the Start license Button
        row = wx.BoxSizer(orient=wx.HORIZONTAL)
        self.start_license_button = wx.Button(self, label="&Upload License")
        self.start_license_button.SetToolTip("Upload the License")
        row.AddStretchSpacer()
        row.Add(self.start_license_button)
        # main.Add(row, border=UNIT // 6)
        main.Add(row, border=UNIT * 6, flag=wx.EXPAND)
        row.AddStretchSpacer()
        # Create a Spacer between Start license button and bottom of dialog
        main.Add(wx.StaticBox(self, size=(1, 20)))
        self.SetSizer(main)
        self.Fit()
        self.start_license_button.Bind(wx.EVT_BUTTON, self.OnStartlicense)
        self.choose_file_button.Bind(wx.EVT_BUTTON, self.OnUpload)
        #self.SetAffirmativeId(wx.ID_OK)

    def refresh_supported_devices(self):
        """
            This method is created to refresh the display Upgrades List Control.
        """
        self.license_list.SetItemCount(len(self.supported_devices))
        # Sort devices by numerical IP address
        self.supported_devices.sort(key=lambda item: item.IP32)
        self.license_list.Update()

    def ValidateLicenseFile(self, model):
        """
        returns a touple (size, date)
            size: a string 'small', 'medium' or 'large'
            date: The date.
        """
        supported = False
        try:
            supported = model.validate_license(model.results, self.license_path)
            print(f"License validation passed: {supported}, "
                  f"path: {self.license_path}")
            return supported
        except Exception as error:
            # traceback.print_exc()
            print(("There was a problem validating the " +
                   "license on device %s, card: %s " % (model.IP, model.name) +
                   "%s, error: %s" % (self.license_path, error)))
        return supported

    def OnUpload(self, event):
        """ Provides a FileDialog so the user chooses a file.
            This function validates looks through our devices list and will
            go through each device's check_license_support method which
            validates if this file fits the description for the license

            This function populates:
                self.supported_devices: with devices that are validated/supported
                self.license_path: The path of the selected file.
                self.license_file: This variable will hold the whole filepath

            This function changes UI:
                chosen_file_label: Displays 'File Chosen: '
                license_chosen_file_label: Displays only the filename
        """
        self.supported_devices = []
        self.chosen_file_label.Show(False)
        self.license_chosen_file_label.Show(False)
        self.license_list.DeleteAllItems()
        self.license_list.Update()
        # otherwise ask the user what new file to open
        with wx.FileDialog(self, "Open license file", wildcard="license files (*)|*",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return     # the user changed their mind
            # Proceed loading the file chosen by the user
            self.license_path = fileDialog.GetPath()
            try:
                self.license_file = open(self.license_path, 'rb')
            except IOError:
                wx.LogError("Cannot open file '%s'." % self.license_path)
                return None
        # Check if the user scanned already
        devices = self.parent.added_devices
        if devices:
            # Show the file Chosen Label
            self.chosen_file_label.Show(True)
            self.license_chosen_file_label.Show(True)
            self.license_chosen_file_label.SetLabelText(str(self.license_path.split('\\')[-1]))
            self.Layout()
            # Check if the devices selected are supported
            for device in devices:
                if device.model:
                    # Check if device has modules (sub-devices)
                    try:
                        # Check if the modules are supported by the file
                        if hasattr(device.model, 'modules'):
                            for module in device.model.modules:
                                print(f"DEBUG: module: {module.name}")
                                # supported Returns ()
                                supported = self.ValidateLicenseFile(module)
                                if supported:
                                    print(f"DEBUG: inside OnUpload supported:{supported}")
                                    module.selected = True
                                    self.supported_devices.append(module)
                                    self.refresh_supported_devices()
                                    self.license_info = supported
                                    print(f"DEBUG: module: {module} supported devices: {self.supported_devices}")
                    except AttributeError as error:
                        # If device has no modules continue
                        print(f"Info: No modules found. Error: {error}")
                    # Check if the device itself can be upgraded using the file
                    supported = self.ValidateLicenseFile(device.model)
                    # supported Returns ()
                    if not supported:
                        continue
                    # If its supported add it to the supported devices
                    device.selected = True
                    self.supported_devices.append(device.model)
                    # Update the virtual list control with the number of devices.
                    self.refresh_supported_devices()
                    # license_info ex (size, date)
                    self.license_info = supported
                    wx.SafeYield()
        else:
            self.ClearAll()
            wx.LogError("There are no devices found. Please scan a subnet")
        # if self.supported_devices is not populated
        if not self.supported_devices:
            self.ClearAll()
            wx.LogError(("There are no devices found which match the license " +
                         "you have chosen license passed. license: " +
                         "%s" % (self.license_path)))

    def GetSelectedDevices(self):
        """
            Sets the number of devices that are selected to be upgraded

            Updates self.num_selected
        """
        num_selected = 0
        # Count how many devices they want to upgrade.
        for device in self.supported_devices:
            if device.selected:
                num_selected += 1
        return num_selected

    def OnStartlicense(self, event):
        """ This method will instantiate the license Upgrade attempt
            process on the cards in self.supported_devices.
        """
        # TODO update this method if needed Get the number of devices
        self.num_selected = self.GetSelectedDevices()
        if self.num_selected < 1:
            wx.MessageDialog(self, "No devices chosen to Upload the license to. " +
                             f"Number of selected devices: {self.num_selected}",
                             caption="No Devices Found!",
                             style=wx.OK | wx.ICON_WARNING).ShowModal()
            return None
        # self.supported devices is a list of DeviceInfo, we need the model.
        licence_device_list = []
        # Go through each supported device
        for device in self.supported_devices:
            # Only upgrade if the device has a selected checkbox
            if device.selected is False:
                continue
            # TODO Create a Thread and run the Upgrade in there...
            # Go into the device model and attempt to upload the license.
            if 'module' in device.GROUP.lower():
                supported = self.ValidateLicenseFile(device)
                print(f"DEBUG: Module supported:{supported}")
            else:
                # Add it into one variable so it can be used later.
                # device = device.model
                supported = self.ValidateLicenseFile(device)
                print(f"DEBUG: Device supported:{supported}")

            # print(f"DEBUG: Supported:{supported}")
            if not supported:
                # Should never reach this statement
                wx.MessageDialog(self, "The File which you want to " +
                                 f"upload: {self.device.model.name}, "
                                 f"file: {self.license_path} "
                                 "is not supported with the chosen device",
                                 style=wx.ICON_ERROR).ShowModal()
                continue
            licence_device_list.append(device.model)
        # dialog containing the IP and the license to be added
        result = DisplayUpgradeInfoDialog(self, self.wxconfig,
                                          licence_device_list).ShowModal()
        # If the user clicks yes on the dialog.
        if int(result) == 5100:
            # Create a list to start polling
            polling = []
            # Start the progress bar
            upload_progress = wx.ProgressDialog("Starting License Upload",
                                                "Validating Version",
                                                parent=None,
                                                style=wx.PD_SMOOTH |
                                                wx.PD_AUTO_HIDE)
            percent = 10
            upload_progress.Update(percent, ("Starting License Upload",
                                                "Validating Version"
                                    f"1/{len(licence_device_list)}"))
            for device in licence_device_list:
                # Only upload if the device has a selected checkbox
                if device.selected is False:
                    continue
                # Start the upload
                thread = UploadThread(device, self.licence_path)
                thread.start()
                polling.append(device)

            # 2) Poll if the device has finished upgrading
            while polling:
                # 2) After each upgrade is completed update the progress bar
                # Start a timout of 8 minutes
                # start_time = time.time()
                # end_time = start_time + 480
                # Check if the upgrade finished
                for device in licence_device_list:
                    time.sleep(10)
                    print(f"DEBUG: Wait 10 seconds DEVICE UPGRADE FINISHED FLAG: {device.flags['upgradeFinished']}, error:{device.flags['upgradeError']}")
                    if device.flags['upgradeFinished'] is True or device.flags['upgradeError']:
                        # Update the progress bar
                        percent += int((len(polling) / licence_device_list) * 100)
                        if percent > 100:
                            percent = 98
                            upload_progress.Update(percent,
                                                    (f"{licence_device_list} "
                                                    "Waiting for devices to "
                                                    "finish upgrading. "
                                                    f"{len(polling)}/"
                                                    f"{len(licence_device_list)}"))
                        if device in polling:
                            # Remove the device from list
                            polling.remove(device)
                        else:
                            print(f"ERROR: Device was not found in polling. {device.IP}")
            # Check if all devices passed.
            upload_progress.Update(100)
            upload_progress.Destroy()
            DisplayUpgradeInfoDialog(self, self.wxconfig, licence_device_list,
                                     first_display=False).ShowModal()
            # try:
            #     model.upload_license(self.license_path, upload_progress)
            #     # Close the progress bar
            #     upload_progress.Update(100, "Succsesfully uploaded Device")
            #     upload_progress.Destroy()
            # except ValueError as error:
            #     # Close the progress bar
            #     upload_progress.Update(100, "Succsesfully uploaded Device")
            #     upload_progress.Destroy()
                # wx.MessageDialog(self, ("Error: Failed to Upload License " +
                #                  f" to Device IP: {model.IP}, name: " +
                #                  f"{model.name} due to error:{error}"),
                #                  style=wx.ICON_ERROR).ShowModal()
            # Check if the upload finished
            # if model.flags['uploadFinished'] is True:
            #     wx.MessageDialog(self, "Succsesfully uploadd Device %s: %s" %
            #                      (model.IP, model.name)).ShowModal()
        else:
            # There was a problem starting the upload.
            print("User decided against starting the upload.")
            return None

    def ClearAll(self):
        """This method will remove everything from the upgradeList Control"""
        self.supported_devices = []
        self.license_list.DeleteAllItems()
        self.license_list.Update()
#end class license Dialog

#
#end class UploadList

class DisplayUpgradeInfoDialog(wx.Dialog):
    """Dialog to show you the devices that you are about to upload as
       well as the devices that are finished uploading.

    Parameters:
        devices: a list of <Model> objects which are set to be upgraded.
        first_display: Determines if you want to see one button or two <Bool>
    .ListCtrl for each device discovered.
    The user data for each item contains the DeviceInfo object.
    """
    def __init__(self, parent: wx.Panel, wxconfig: wx.ConfigBase, devices,
                 first_display=True):
        wx.Dialog.__init__(self, parent, title="Upgrade Information",
                           style=wx.DEFAULT_DIALOG_STYLE | wx.RESIZE_BORDER,
                           size=(1000, 2000))
        # self.SetSize((1000,1000))
        self.SetMinSize((1000, 450))
        self.parent: wx.Panel = parent
        self.wxconfig: wx.ConfigBase = wxconfig
        self.user_result = 0
        self.selected_devices = devices
        self.firmware_info = self.parent.firmware_info
        # If the device is supported it provides some information on the firmware.
        title_font = wx.Font(18, wx.DECORATIVE, wx.ITALIC, wx.NORMAL)
        # Initialize UI -----------------------------------------------
        UNIT = max(int(wx.SystemSettings.GetMetric(wx.SYS_ICON_X) * 1.125), 36)
        ACVR = wx.ALIGN_CENTER_VERTICAL | wx.ALL
        main = wx.BoxSizer(orient=wx.VERTICAL)

        # Create a Spacer between Top of the dialog and Title
        main.Add(wx.StaticBox(self, size=(1, 10)))
        # Title Row
        # upgrade_title_row = wx.BoxSizer(orient=wx.HORIZONTAL)
        upgrade_main_label = wx.StaticText(self, label="Upgrade Information",
                                           style=wx.ALIGN_CENTER)
        upgrade_main_label.SetFont(title_font)
        main.Add(upgrade_main_label, flag=wx.EXPAND)
        # Add Upgrade list
        # self.upgrade_list = wx.BoxSizer(orient=wx.HORIZONTAL)
        # Add a Vertical Spacer
        main.Add(wx.StaticBox(self, size=(1, 20)))
        # Create the Supported Devices List.
        row = wx.BoxSizer(orient=wx.HORIZONTAL)
        row.AddStretchSpacer()
        self.display_upgrade_list = DisplayUpgradeList(self, devices)
        row.Add(self.display_upgrade_list)
        row.AddStretchSpacer()
        main.Add(row, flag=wx.EXPAND)

        # Create a Spacer between Upgrade CtrList and Start upgrade button
        main.Add(wx.StaticBox(self, size=(1, 20)))
        # Add the Start Upgrade Button
        row = wx.BoxSizer(orient=wx.HORIZONTAL)
        stop_button_text = "&Close"
        # If its the first popup we have to change the
        if first_display:
            # Add the Warning Message
            upgrade_main_label = wx.StaticText(self,
                                               label=("Are you sure you " +
                                                      "want to upgrade using " +
                                                      "the above firmware?"),
                                                      style=wx.ALIGN_CENTER)
            upgrade_main_label.SetForegroundColour((178, 34, 34))
            main.Add(upgrade_main_label, flag=wx.EXPAND)
            # Add the Stop upgrade button
            self.start_upgrade_button = wx.Button(self, label="&Start Upgrade")
            self.start_upgrade_button.SetToolTip("Start the Upgrade")
            row.AddStretchSpacer()
            row.Add(self.start_upgrade_button)
            row.AddStretchSpacer()
            # Set the button text at the end to be close instead of cancel
            stop_button_text = "&Cancel Upgrade"
        self.stop_upgrade_button = wx.Button(self, label=stop_button_text)
        self.stop_upgrade_button.SetToolTip("Cancel the Upgrade")
        row.AddStretchSpacer()
        row.Add(self.stop_upgrade_button)
        row.AddStretchSpacer()
        # main.Add(row, border=UNIT // 6)
        main.Add(row, border=UNIT * 6, flag=wx.EXPAND)
        # Create a Spacer between Start upgrade button and bottom of dialog
        self.SetSizer(main)
        self.Fit()

        if first_display:
            # returns 5100 if start button is pressed
            self.start_upgrade_button.Bind(wx.EVT_BUTTON, self.OnStartUpgrade)
        # returns something else if this button is pressed.
        self.stop_upgrade_button.Bind(wx.EVT_BUTTON, self.OnClose)

    def refresh_supported_devices(self):
        """
            This method is created to refresh the Upgrade List Control.
        """
        self.display_upgrade_list.SetItemCount(len(self.selected_devices))
        # Sort devices by numerical IP address
        self.selected_devices.sort(key=lambda item: item.IP32)
        self.display_upgrade_list.Update()

    def OnStartUpgrade(self, event):
        """
            The user wants to start the upgrade. Close the dialog and
            return 5100

            Note: 5100 is the 'yes' return statement for wx.MessageDialog.
        """
        self.user_result = 5100
        print("DEBUG: User returned yes 5100")
        # Close the dialog
        self.SetReturnCode(5100)
        self.Destroy()

    def OnClose(self, event):
        """
            The user wants to cancel the upgrade. Close the dialog
            return anything other than 5100
        """
        self.user_result = 0
        print("DEBUG: User returned no 0")
        self.SetReturnCode(0)
        # Close the dialog
        self.Destroy()

# End of DisplayUpgradeInfoDialog

class DisplayUpgradeList(wx.ListCtrl):
    """List control to show you the devices that you are about to upload as
       well as the devices that are finished uploading.

    Expects Parent to have values populated for:
        self.parent.firmware_info
        self.parent.selected_devices
        self

    self.devices = List of DeviceInfo class to show.
    NOTE: device needs to have device.selected to work.

    If OnUpload (OnOpen) has been run these will be populated.
        self.chosen_filename: The name of the chosen file
        self.image_file: Contains the file object
    """
    COLUMNS = (("Status", wx.LIST_FORMAT_CENTER, 1.5),
               ("IP", wx.LIST_FORMAT_LEFT, 3.6),
               ("Device Name", wx.LIST_FORMAT_LEFT, 7.3),
               ("Details", wx.LIST_FORMAT_CENTRE, 16.0))

    def __init__(self, parent, devices: dict, menu=None, style=0):
        # Call the original constructor to do its job. Force style
        style |= wx.LC_REPORT | wx.LC_VIRTUAL
        wx.ListCtrl.__init__(self, parent, style=style, size=(1000, 250))
        # Assign our parent book and figure out our page number.
        self.parent: wx.Window = parent             # Parent window
        self.menu: wx.Menu = menu                   # Popup context menu.
        # This is all devices in our list control.
        self.selected_devices: typing.Dict[int, DeviceInfo] = devices
        # Setup the item count #NEEDED FOR ListCtrl
        self.SetItemCount(len(self.selected_devices))
        self.selected_devices.sort(key=lambda item: item.IP32)
        self.Update()
        # Add standard columns
        UNIT = max(int(wx.SystemSettings.GetMetric(wx.SYS_ICON_X) * 1.125), 36)
        for title, alignment, width in self.COLUMNS:
            self.AppendColumn(title, format=alignment, width=int(width * UNIT))
        # self.Bind(wx.EVT_RIGHT_DOWN, self.OnRightDown)
        # Create Checkbox images
        self.il = wx.ImageList(16, 16)
        self.check_image = self.il.Add(self.CreateBitmap(wx.CONTROL_CHECKED, (16, 16)))
        self.uncheck_image = self.il.Add(self.CreateBitmap(0, (16, 16)))
        self.SetImageList(self.il, wx.IMAGE_LIST_SMALL)

    def CreateBitmap(self, flag=0, size=(16, 16)):
        """Create a bitmap of the platforms native checkbox. The flag
        is used to determine the checkboxes state (see wx.CONTROL_*)
        """
        bmp = wx.Bitmap(*size)
        dc = wx.MemoryDC(bmp)
        dc.SetBackground(wx.WHITE_BRUSH)
        dc.Clear()
        wx.RendererNative.Get().DrawCheckBox(self, dc,
                                             (0, 0, size[0], size[1]), flag)
        dc.SelectObject(wx.NullBitmap)
        return bmp

    def OnGetItemAttr(self, item_num):
        """Return the  item number (row) and column number.
           NOTE: With this method we are passed the device.
           Not the DeviceInfo object.
        """
        if item_num >= 0 and item_num < len(self.selected_devices):
            device_obj = self.selected_devices[item_num]
            upgrade_details = device_obj.flags.get('upgradeError')
            upgrade_finished = device_obj.flags.get('upgradeFinished')
            if upgrade_details:
                # If there was a failure with the upgrade. Change color
                self.SetItemBackgroundColour(item_num,
                                             wx.Colour(247, 103, 101))
                # wx.Colour(247, 103, 101)
            elif upgrade_finished:
                # If the upgrade passed
                self.SetItemBackgroundColour(item_num,
                                             wx.Colour(38, 156, 98))
                #  wx.Colour(38, 156, 98)
            else:
                return None
        # Base class returns None.
        return None

    def OnGetItemText(self, item_num, column):
        """Return text for the given item number (row) and column number.
           NOTE: With this method we are passed the device.
           Not the DeviceInfo object.
        """
        if item_num >= 0 and item_num < len(self.selected_devices):
            device_obj = self.selected_devices[item_num]

            if column == 0:
                # Uses an image instead.
                return ""
            if column == 1:         # Card IP
                return device_obj.IP
            if column == 2:         # Card Name/Description of the card
                # If the device is a submodule, get that models name
                try:
                    if "module" in device_obj.GROUP.lower():
                        return device_obj.name
                    else:
                        return device_obj.name
                except AttributeError:
                    return device_obj.name
            if column == 3:         # Date
                # Get the information about the upgrade
                upgrade_details = device_obj.flags.get('upgradeError')
                upgrade_finished = device_obj.flags.get('upgradeFinished')
                if upgrade_details:
                    # If there was a failure with the upgrade. Change color
                    self.SetItemBackgroundColour(item_num,
                                                 wx.Colour(247, 103, 101))
                    # wx.Colour(247, 103, 101)
                    return upgrade_details
                elif upgrade_finished:
                    # If the upgrade passed
                    self.SetItemBackgroundColour(item_num,
                                                 wx.GREEN)
                    #  wx.Colour(38, 156, 98)
                    return f"Successful Upgrade on {device_obj.name}."
                else:
                    return f"About to be upgraded to {self.parent.firmware_info}"
        # Unhandled row/column. Return blank string.
        return ""

    def OnGetItemImage(self, item):
        """Return the item checkbox image for the given integer item index."""
        devices = self.selected_devices
        if item >= 0 and item < len(devices):
            device = devices[item]
            upgrade_error = device.flags.get('upgradeError')
            if upgrade_error:
                return self.uncheck_image
            elif device.flags.get('upgradeFinished'):
                return self.check_image
            else:
                return self.uncheck_image
        # Outside of list! Use no image.
        return -1
#end class DisplayUpgradeList

class UpgradeThread(threading.Thread):
    """A Thread to perform device upgrade on a list of devices (Models)"""
    def __init__(self, device, firmware_path, firmware_file,
                 supported_devices, name="Upgrade Thread", **kwargs):
        """
        Parameters:
            firmware_path: Path to firmware that user chose
            firmware_file: Opened Firmware in bytes
            supported_devices: List of supported devices.
        """
        # Initialize Thread.
        super(UpgradeThread, self).__init__(name=name, **kwargs)
        self.device = device
        self.firmware_path = firmware_path
        self.firmware_file = firmware_file
        # Create new HTTP, SNMP, and NBT threads for this scan.
        self.http_thread = ahttp.start()
        self.snmp_thread = asnmp.start()
        # self.polling = supported_devices
        self.discovered = collections.deque()
        self.end_event = threading.Event()
        self.percent = 0
        self.status = "Initializing"

    def run(self):
        """This is the method that is run in a separate thread.
            This method will start a thread which will upgrade the device.
        """
        # 1) Upgrade the device
        self.status = f"Upgrading {self.device.IP}"
        try:
            self.device.upgrade(self.firmware_path)
        except ValueError as error:
            self.status = "Exception %s" % str(error)
            # return ("Error: Failed to Upgraded Device "
            #                  f"IP: {self.device.IP}, name: "
            #                  f"{self.device.name} due to error:{error}")
        else:
            self.status = "Finished"
        # 3) Stop all threads.
        self.http_thread.stop(block=False)
        self.snmp_thread.stop(block=False)

    def stop(self, block=True):
        """Stop the thread, optionally blocking until fully stopped."""
        self.end_event.set()
        if block is True:
            self.join()
#end class UpgradeThread(threading.Thread)

class UploadThread(threading.Thread):
    """A Thread to perform device upload on a list of devices (Models)"""
    def __init__(self, device, licence_path,
                 name="Upload Thread", **kwargs):
        """
        Parameters:
            licence_path: Path to firmware that user chose
            licence_device_list: List of supported devices.
        """
        # Initialize Thread.
        super(UploadThread, self).__init__(name=name, **kwargs)
        self.device = device
        self.licence_path = licence_path
        # Create new HTTP, SNMP, and NBT threads for this scan.
        self.http_thread = ahttp.start()
        self.snmp_thread = asnmp.start()
        # self.polling = supported_devices
        self.discovered = collections.deque()
        self.end_event = threading.Event()
        self.percent = 0
        self.status = "Initializing"

    def run(self):
        """This is the method that is run in a separate thread.
            This method will start a thread which will upload the device.
        """
        # 1) Upload the device
        self.status = f"Uploading License to {self.device.IP}"
        try:
            self.device.upload_license(self.license_path)
        except ValueError as error:
            #TODO No check for exception in threads.
            self.status = "Exception %s" % str(error)
        else:
            self.status = "Finished"
        # 3) Stop all threads.
        self.http_thread.stop(block=False)
        self.snmp_thread.stop(block=False)

    def stop(self, block=True):
        """Stop the thread, optionally blocking until fully stopped."""
        self.end_event.set()
        if block is True:
            self.join()
#end class UploadThread(threading.Thread)

# wxPython Entrypoint ----------------------------------------------------
# Create Main Frame (and all sub-windows) and run wxPython event loop.
if __name__ == "__main__":
    import multiprocessing
    ## This line is only required for multiprocessing support with pyinstaller.
    ## https://github.com/pyinstaller/pyinstaller/wiki/Recipe-Multiprocessing
    multiprocessing.freeze_support()
    # Create Main Frame and application GUI layout.
    frame = AppFrame(DEBUG)
    # Run the message pump forever until a QUIT message is received.
    app.MainLoop()
